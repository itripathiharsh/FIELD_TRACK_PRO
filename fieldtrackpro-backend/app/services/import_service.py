"""
Import Service: the full Excel/MIS import pipeline for FieldTrack.

Supports:
1. Real Employee Master Imports (Employee ID, Name, DOB, Working Profile, Address, Mail, Phone, CUG)
   - Provisions User accounts with secure hashed temporary passwords
   - Creates/Updates Employee records matching on unique employee_code
   - Generates credential/onboarding Excel spreadsheets
2. Real DMS & Combined BI Outlet / Financial Imports
   - Hierarchy: Zone (Territory) -> Area -> Outlet (Customer with unique dms_code)
   - FOS name matching engine with alias resolution & exception handling
   - Employee ↔ Outlet direct assignments
   - Non-destructive OutletFinancialSnapshot with Sales, Collection, Market OS & 7 Ageing Buckets
3. Legacy Invoice / Payment Imports
"""
from __future__ import annotations

import io
import re
import uuid
import secrets
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, or_, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions.custom import BaseAPIException
from app.models.customer import Customer
from app.models.employee import Employee
from app.models.import_batch import ImportBatch, ImportStatus
from app.models.invoice import Invoice, InvoiceSource
from app.models.payment import Payment, PaymentMethod, PaymentSource, PaymentStatus
from app.models.territory import Territory
from app.models.area import Area
from app.models.user import User, Role
from app.models.employee_customer_assignment import EmployeeCustomerAssignment
from app.models.fos_mapping import FOSEmployeeMapping
from app.models.outlet_financial_snapshot import OutletFinancialSnapshot
from app.core.security import hash_password

MAX_PREVIEW_ROWS = 20
MAX_FILE_SIZE_BYTES = 15 * 1024 * 1024

# ---------------------------------------------------------------------------
# Target field registry
# ---------------------------------------------------------------------------
TARGET_FIELDS: dict[str, dict[str, Any]] = {
    # Geographic hierarchy
    "zone_name":                    {"label": "Zone / Territory",          "required": False, "aliases": ["zone", "zone_name", "territory", "region", "town", "district"]},
    "area_name":                    {"label": "Area",                      "required": False, "aliases": ["area", "area_name", "locality", "sub-zone", "market area", "location"]},
    
    # Employee Master
    "employee_code":                {"label": "Employee Code / ID",        "required": False, "aliases": ["employee id", "employee code", "emp code", "emp id", "fos code", "sr. no.", "sr no", "sr_no", "s.no", "s no"]},
    "employee_name":                {"label": "Employee Name",             "required": False, "aliases": ["employee name", "employee", "salesperson", "rep name", "name", "staff name", "executive name"]},
    "employee_email":               {"label": "Employee Mail ID",          "required": False, "aliases": ["mail id", "email", "mail", "e-mail", "employee email", "email id", "email address"]},
    "employee_phone":               {"label": "Employee Phone No.",        "required": False, "aliases": ["phone no.", "phone no", "phone", "mobile no", "mobile", "mobile phone", "contact no", "contact number"]},
    "employee_working_profile":     {"label": "Working Profile",           "required": False, "aliases": ["working profile", "profile", "designation", "job title", "role in company", "department"]},
    "employee_cug":                 {"label": "CUG Number",                "required": False, "aliases": ["cug", "cug no", "cug number", "sim cug", "official cug"]},
    "employee_dob":                 {"label": "Date of Birth",             "required": False, "aliases": ["date of birth", "dob", "birth date"]},
    "employee_address":             {"label": "Employee Address",          "required": False, "aliases": ["employee address", "emp address", "residence address", "address", "permanent address"]},
    "employee_father_name":         {"label": "Father Name",               "required": False, "aliases": ["father name", "father's name"]},
    "employee_mother_name":         {"label": "Mother Name",               "required": False, "aliases": ["mother name", "mother's name"]},
    "employee_aadhaar":             {"label": "Aadhar No.",                "required": False, "aliases": ["aadhar no.", "aadhar no", "aadhar", "aadhaar no", "aadhaar", "aadhaar number"]},
    "employee_pan":                 {"label": "PAN No.",                   "required": False, "aliases": ["pan no.", "pan no", "pan", "pan number"]},
    "employee_app_role":            {"label": "Application Role",          "required": False, "aliases": ["app role", "application role", "system role", "user role"]},

    # Outlet / Customer Master
    "outlet_code":                  {"label": "Outlet Code",               "required": False, "aliases": ["outlet code", "outlet_code", "outlet id", "customer id", "retailer code", "client id", "party code", "customer code"]},
    "dms_code":                     {"label": "DMS Code",                  "required": False, "aliases": ["dms code", "dms_code", "dms", "dmscode"]},
    "outlet_name":                  {"label": "Outlet / Customer Name",    "required": False, "aliases": ["outlet_name", "outlet name", "outlet", "retailer", "retailer name", "customer name", "shop name", "party name", "account name"]},
    "outlet_address":               {"label": "Outlet Address",            "required": False, "aliases": ["shop address", "outlet address", "location", "store address"]},
    "outlet_contact_number":        {"label": "Outlet Contact Number",     "required": False, "aliases": ["contact", "contact number", "outlet phone", "retailer phone"]},
    "outlet_contact_person":        {"label": "Outlet Contact Person",     "required": False, "aliases": ["contact person", "owner name", "proprietor", "key contact"]},
    "outlet_latitude":              {"label": "Outlet Latitude",           "required": False, "aliases": ["latitude", "lat"]},
    "outlet_longitude":             {"label": "Outlet Longitude",          "required": False, "aliases": ["longitude", "lng", "long"]},
    "outlet_geofence_radius":       {"label": "Geofence Radius (m)",       "required": False, "aliases": ["geofence radius", "radius", "radius_m"]},
    "fos_name":                     {"label": "FOS Name / Sales Rep",      "required": False, "aliases": ["fos_name", "fos name", "fos", "assigned fos", "sales rep name", "sales rep", "rep name", "sales executive", "executive"]},
    
    # Geographic hierarchy
    "territory_name":               {"label": "Territory Name",            "required": False, "aliases": ["territory", "territory name", "territory_name"]},
    "zone_name":                    {"label": "Zone / Territory",          "required": False, "aliases": ["zone", "zone_name", "region", "town", "district"]},
    "area_name":                    {"label": "Area",                      "required": False, "aliases": ["area", "area_name", "locality", "sub-zone", "market area", "location"]},

    # Financial / Business BI
    "brand":                        {"label": "Brand Name",                "required": False, "aliases": ["brand name", "division", "product line"]},
    "invoice_brand":                {"label": "Invoice Brand",             "required": False, "aliases": ["brand", "invoice brand", "invoice_brand"]},
    "sales_amount":                 {"label": "Sales Amount",              "required": False, "aliases": ["sales", "sales amount", "total sales", "billed amount", "sale", "net sales"]},
    "collection_amount":            {"label": "Collection Amount",         "required": False, "aliases": ["collection", "collection amount", "total collection", "received amount", "collected"]},
    "market_outstanding":           {"label": "Market Outstanding (OS)",   "required": False, "aliases": ["market_os", "market os", "market outstanding", "total os", "outstanding", "balance", "os amount", "os", "market_os (inr)", "total outstanding"]},
    "bucket_lt_15":                 {"label": "Ageing < 15 Days",          "required": False, "aliases": ["<15 days", "(< 15 days )", "< 15 days", "lt 15 days", "0-15 days", "<15", "0 to 15 days", "upto 15 days"]},
    "bucket_15_30":                 {"label": "Ageing 15 - 30 Days",       "required": False, "aliases": ["15-30 days", "15 to 30 days", "15 - 30 days", "15-30", "16 to 30 days"]},
    "bucket_30_45":                 {"label": "Ageing 30 - 45 Days",       "required": False, "aliases": ["30-45 days", "30 to 45 days", "30 - 45 days", "30-45", "31 to 45 days"]},
    "bucket_45_60":                 {"label": "Ageing 45 - 60 Days",       "required": False, "aliases": ["45-60 days", "45 to 60 days", "45 - 60 days", "45-60", "46 to 60 days"]},
    "bucket_60_75":                 {"label": "Ageing 60 - 75 Days",       "required": False, "aliases": ["60-75 days", "60 to 75 days", "60 - 75 days", "60-75", "61 to 75 days"]},
    "bucket_75_90":                 {"label": "Ageing 75 - 90 Days",       "required": False, "aliases": ["75-90 days", "75 to 90 days", "75 - 90 days", "75-90", "76 to 90 days"]},
    "bucket_gt_90":                 {"label": "Ageing > 90 Days",          "required": False, "aliases": [">90 days", "(> 90 days )", "> 90 days", "gt 90 days", "above 90 days", ">90", "above 90", "> 90 days "]},
    
    # Legacy Invoice & Payment fields
    "invoice_number":               {"label": "Invoice Number",            "required": False, "aliases": ["invoice no", "invoice number", "bill no", "voucher no"]},
    "invoice_date":                 {"label": "Invoice Date",              "required": False, "aliases": ["invoice date", "bill date", "voucher date"]},
    "invoice_amount":               {"label": "Invoice Amount",            "required": False, "aliases": ["invoice amount", "bill amount"]},
    "payment_amount":               {"label": "Payment Amount",            "required": False, "aliases": ["payment amount", "paid amount"]},
    "payment_date":                 {"label": "Payment Date",              "required": False, "aliases": ["payment date", "collection date"]},
    "payment_method":               {"label": "Payment Method",            "required": False, "aliases": ["payment method", "mode", "payment mode"]},
    "payment_reference":            {"label": "Payment Reference / UTR",   "required": False, "aliases": ["utr", "reference", "cheque no", "cheque number", "ref no"]},
}

OUTLET_STRATEGY_CODE = "outlet_code"
OUTLET_STRATEGY_NAME_TERRITORY = "name_and_territory"


def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _validate_upload(file_bytes: bytes, filename: str) -> None:
    if len(file_bytes) == 0:
        raise BaseAPIException(status_code=400, detail="The uploaded file is empty", error_code="IMPORT_EMPTY_FILE")
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        raise BaseAPIException(
            status_code=400,
            detail=f"File exceeds maximum allowed size of {MAX_FILE_SIZE_BYTES // (1024 * 1024)}MB",
            error_code="IMPORT_FILE_TOO_LARGE",
        )
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise BaseAPIException(
            status_code=415,
            detail="Only Excel spreadsheets (.xlsx, .xls) are supported",
            error_code="IMPORT_INVALID_EXTENSION",
        )


def detect_sheet_type(columns: list[str]) -> str:
    norm_cols = [_normalize(c) for c in columns]
    joined = " ".join(norm_cols)
    if any(k in joined for k in ["working profile", "cug", "father name", "aadhar", "pan no", "employee code"]):
        return "employee_master"
    if any(k in joined for k in ["dms code", "market os", "market_os", "<15 days", "< 15 days", "fos name", "outlet_name"]):
        return "dms_outlet_financial"
    if any(k in joined for k in ["sales", "collection", "outstanding", "brand"]):
        return "combined_bi"
    return "generic"


def suggest_column_mapping(columns: list[str]) -> dict[str, Optional[str]]:
    suggestion: dict[str, Optional[str]] = {}
    used_targets: set[str] = set()
    for col in columns:
        norm = _normalize(col)
        matched_key = None
        for key, spec in TARGET_FIELDS.items():
            if key in used_targets:
                continue
            names = [key.replace("_", " ")] + spec["aliases"]
            if norm in (_normalize(n) for n in names):
                matched_key = key
                break
        if matched_key:
            used_targets.add(matched_key)
        suggestion[col] = matched_key
    return suggestion


def find_header_row_and_data(rows: list[tuple[Any, ...]]) -> tuple[int, list[str], list[list[Any]]]:
    """
    Scans the top 25 rows of a spreadsheet to detect the actual table header row.
    Skips title banners, metadata blocks, and summary/total rows.
    Returns (header_row_index (0-indexed), header_column_names, clean_data_rows).
    """
    if not rows:
        return 0, [], []
    
    best_row_idx = 0
    best_score = -1
    best_headers: list[str] = []

    for idx, row in enumerate(rows[:25]):
        if not row:
            continue
        cells = [c for c in row if c is not None and str(c).strip()]
        if not cells:
            continue
        
        first_cell_str = str(row[0]).strip().lower() if row[0] is not None else ""
        if first_cell_str in ["total", "subtotal", "grand total"]:
            continue
        
        # Check how many cells match known target aliases
        match_count = 0
        used_keys: set[str] = set()
        for cell in row:
            if cell is None:
                continue
            norm = _normalize(str(cell))
            for key, spec in TARGET_FIELDS.items():
                if key in used_keys:
                    continue
                names = [key.replace("_", " ")] + spec["aliases"]
                if norm in (_normalize(n) for n in names):
                    match_count += 1
                    used_keys.add(key)
                    break
        
        # Score calculation: strong boost for target matches, slight boost for unique text strings, penalty for long paragraphs
        is_title = any(len(str(c).strip()) > 45 for c in cells)
        score = (match_count * 15) + len(cells) - (50 if is_title else 0)
        
        if match_count >= 2 and score > best_score:
            best_score = score
            best_row_idx = idx
            best_headers = [str(c).strip() if (c is not None and str(c).strip()) else f"(blank column {i+1})" for i, c in enumerate(row)]

    # Fallback if no target field cluster found
    if best_score < 10:
        for idx, row in enumerate(rows[:10]):
            cells = [c for c in row if c is not None and str(c).strip()]
            if len(cells) >= 2 and not any(str(c).lower().startswith("total") for c in cells):
                best_row_idx = idx
                best_headers = [str(c).strip() if (c is not None and str(c).strip()) else f"(blank column {i+1})" for i, c in enumerate(row)]
                break
        if not best_headers and rows:
            best_row_idx = 0
            best_headers = [str(c).strip() if (c is not None and str(c).strip()) else f"(blank column {i+1})" for i, c in enumerate(rows[0])]

    raw_data_rows = rows[best_row_idx + 1 :]
    # Clean data rows: filter out trailing total/summary rows and completely empty rows
    clean_data_rows: list[list[Any]] = []
    for r in raw_data_rows:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        first_val = str(r[0]).strip().lower() if r[0] is not None else ""
        if first_val in ["total", "subtotal", "grand total"]:
            continue
        if len(r) > 1 and str(r[1]).strip().lower() in ["total", "grand total"]:
            continue
        clean_data_rows.append(list(r))

    return best_row_idx, best_headers, clean_data_rows


def check_mapping_confidence(detected_type: str, mapping: dict[str, Optional[str]]) -> tuple[bool, list[str]]:
    mapped_targets = {v for v in mapping.values() if v is not None}
    missing: list[str] = []
    if detected_type in ["dms_outlet_financial", "combined_bi"]:
        if "dms_code" not in mapped_targets and "outlet_name" not in mapped_targets:
            missing.append("DMS Code or Outlet Name")
        if "market_outstanding" not in mapped_targets and not ("sales_amount" in mapped_targets and "collection_amount" in mapped_targets):
            missing.append("Market Outstanding or Sales/Collection")
    elif detected_type == "employee_master":
        if "employee_code" not in mapped_targets:
            missing.append("Employee Code")
        if "employee_name" not in mapped_targets:
            missing.append("Employee Name")
    elif detected_type == "invoice_payment":
        if "invoice_number" not in mapped_targets and "payment_amount" not in mapped_targets:
            missing.append("Invoice Number or Payment Amount")
    
    is_confident = len(missing) == 0 and len(mapped_targets) >= 2
    return is_confident, missing


def preview_excel_file(file_bytes: bytes, filename: str, sheet_name: str | None = None) -> dict:
    _validate_upload(file_bytes, filename)
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        selected_sheet = None
        if sheet_name and sheet_name in workbook.sheetnames:
            selected_sheet = workbook[sheet_name]
        else:
            # Score each sheet to automatically pick the primary data sheet
            best_sheet = None
            best_sheet_score = -1
            for sname in workbook.sheetnames:
                ws = workbook[sname]
                s_rows = list(ws.iter_rows(values_only=True))
                if not s_rows:
                    continue
                h_idx, cols, d_rows = find_header_row_and_data(s_rows)
                suggested = suggest_column_mapping(cols)
                matched_cnt = sum(1 for v in suggested.values() if v is not None)
                # Boost if sheet name is 'Raw Data' or 'Data' or 'Outlets'
                sname_lower = sname.lower().strip()
                name_boost = 50 if sname_lower in ["raw data", "data", "outlets", "sheet1"] else 0
                sheet_score = (matched_cnt * 20) + min(len(d_rows), 500) + name_boost
                if sheet_score > best_sheet_score:
                    best_sheet_score = sheet_score
                    best_sheet = ws
            selected_sheet = best_sheet if best_sheet is not None else workbook.active
    except Exception as err:
        raise BaseAPIException(
            status_code=400,
            detail=f"Could not read Excel workbook: {err}",
            error_code="IMPORT_UNREADABLE_FILE",
        )

    rows = list(selected_sheet.iter_rows(values_only=True))
    all_sheets = workbook.sheetnames
    sheet_title = selected_sheet.title
    workbook.close()

    if not rows:
        raise BaseAPIException(status_code=400, detail="The file contains no data rows", error_code="IMPORT_NO_ROWS")

    header_row_idx, columns, data_rows = find_header_row_and_data(rows)
    sample_rows = [[("" if c is None else str(c)) for c in r] for r in data_rows[:MAX_PREVIEW_ROWS]]
    total_data_rows = len(data_rows)

    detected_type = detect_sheet_type(columns)
    suggested = suggest_column_mapping(columns)
    is_confident, missing_crit = check_mapping_confidence(detected_type, suggested)

    matched_count = sum(1 for v in suggested.values() if v is not None)
    total_columns_count = len(columns)

    return {
        "sheet_name": sheet_title,
        "all_sheets": all_sheets,
        "columns": columns,
        "sample_rows": sample_rows,
        "total_data_rows": total_data_rows,
        "truncated": total_data_rows > MAX_PREVIEW_ROWS,
        "detected_type": detected_type,
        "suggested_mapping": suggested,
        "target_fields": TARGET_FIELDS,
        "unmatched_fos_names": [],
        "header_row_index": header_row_idx + 1,
        "is_confident": is_confident,
        "matched_columns_count": matched_count,
        "total_columns_count": total_columns_count,
        "detected_entity_count": total_data_rows,
    }


# ---------------------------------------------------------------------------
# Validation & Execution Planning
# ---------------------------------------------------------------------------

def _parse_decimal(val: Any) -> Decimal:
    if val is None:
        return Decimal("0.00")
    s = str(val).strip().replace(",", "").replace("₹", "").replace("$", "")
    if not s or s == "-":
        return Decimal("0.00")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _parse_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


async def create_import_batch(
    file_bytes: bytes,
    filename: str,
    sheet_name: str,
    column_mapping: dict[str, str],
    outlet_match_strategy: str,
    allow_generated_invoice_numbers: bool,
    current_user: User,
    session: AsyncSession,
    fos_mapping_overrides: dict[str, uuid.UUID] | None = None,
) -> ImportBatch:
    """
    Parses and validates the Excel upload. Identifies whether it's Employee Master
    or DMS/BI Outlet/Financial data, resolves Zone/Area/Employee identities, and builds
    a complete execution plan in `ImportBatch`.
    """
    _validate_upload(file_bytes, filename)
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active

    # Fetch existing reference caches
    emp_res = await session.execute(select(Employee))
    all_employees = emp_res.scalars().all()
    emp_by_code = {e.employee_code.lower(): e for e in all_employees if e.employee_code}
    emp_by_name = {e.full_name.strip().lower(): e for e in all_employees if e.full_name}

    fos_map_res = await session.execute(select(FOSEmployeeMapping))
    all_fos_mappings = fos_map_res.scalars().all()
    fos_map = {m.raw_fos_name.strip().lower(): m.employee_id for m in all_fos_mappings}
    if fos_mapping_overrides:
        for k, v in fos_mapping_overrides.items():
            fos_map[k.strip().lower()] = v

    cust_res = await session.execute(select(Customer))
    all_customers = cust_res.scalars().all()
    cust_by_code = {c.outlet_code.lower(): c for c in all_customers if c.outlet_code}

    terr_res = await session.execute(select(Territory))
    all_territories = terr_res.scalars().all()
    terr_by_name = {t.name.strip().lower(): t for t in all_territories if t.name}

    area_res = await session.execute(select(Area))
    all_areas = area_res.scalars().all()
    area_by_name_and_terr = {(a.name.strip().lower(), a.territory_id): a for a in all_areas if a.name}

    all_sheet_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not all_sheet_rows:
        raise BaseAPIException(status_code=400, detail="Sheet contains no data rows", error_code="IMPORT_NO_ROWS")

    header_row_idx, header_row, data_rows = find_header_row_and_data(all_sheet_rows)

    col_index_to_target = {}
    for idx, col_name in enumerate(header_row):
        target = column_mapping.get(col_name)
        if target:
            col_index_to_target[idx] = target

    detected_type = detect_sheet_type(list(column_mapping.values()))

    plan_rows: list[dict] = []
    error_report: list[dict] = []
    unmatched_fos_names: set[str] = set()
    rows_created = 0
    rows_updated = 0
    rows_skipped = 0
    rows_error = 0

    row_num = header_row_idx + 1
    for row in data_rows:
        row_num += 1
        if not any(c is not None and str(c).strip() for c in row):
            continue
        first_val = str(row[0]).strip().lower() if row[0] is not None else ""
        if first_val in ["total", "subtotal", "grand total"]:
            continue

        row_data: dict[str, Any] = {}
        for idx, target in col_index_to_target.items():
            if idx < len(row):
                row_data[target] = row[idx]

        if not row_data:
            continue

        # Process Row based on type
        if detected_type == "employee_master" or "employee_code" in row_data or "employee_working_profile" in row_data:
            emp_code = str(row_data.get("employee_code") or "").strip()
            emp_name = str(row_data.get("employee_name") or "").strip()
            if not emp_code or not emp_name:
                error_report.append({"row": row_num, "error": "Missing Employee ID or Employee Name", "suggested_fix": "Provide Employee ID and Name"})
                rows_error += 1
                continue

            email = str(row_data.get("employee_email") or "").strip()
            if not email or "@" not in email:
                email = f"emp{emp_code}@client.fieldtrack.internal"

            mobile = str(row_data.get("employee_phone") or "").strip()
            cug = str(row_data.get("employee_cug") or "").strip()
            working_profile = str(row_data.get("employee_working_profile") or "FOS").strip()
            app_role = str(row_data.get("employee_app_role") or "EMPLOYEE").strip().upper()
            if app_role not in ["ADMIN", "EMPLOYEE"]:
                app_role = "EMPLOYEE"

            dob = _parse_date(row_data.get("employee_dob"))
            address = str(row_data.get("employee_address") or "").strip()

            is_existing = emp_code.lower() in emp_by_code
            if is_existing:
                rows_updated += 1
            else:
                rows_created += 1

            # Generate temporary password for onboarding credential sheet
            temp_pass = f"EmpPass@{secrets.token_hex(3)}!"

            plan_rows.append({
                "type": "employee",
                "employee_code": emp_code,
                "full_name": emp_name,
                "email": email,
                "mobile_phone": mobile,
                "cug": cug,
                "working_profile": working_profile,
                "application_role": app_role,
                "date_of_birth": dob.isoformat() if dob else None,
                "address": address,
                "temp_password": temp_pass,
                "is_existing": is_existing,
            })

        else:
            # DMS / BI Outlet & Financial Row
            dms_code = str(row_data.get("dms_code") or "").strip()
            outlet_name = str(row_data.get("outlet_name") or "").strip()
            if not dms_code and not outlet_name:
                rows_skipped += 1
                continue

            if not dms_code:
                error_report.append({"row": row_num, "error": "Missing DMS Code", "suggested_fix": "Provide a unique DMS Code for the outlet"})
                rows_error += 1
                continue

            zone_name = str(row_data.get("zone_name") or "General Zone").strip()
            area_name = str(row_data.get("area_name") or "").strip()
            if not area_name or area_name == "-":
                area_name = "General Area"

            raw_fos = str(row_data.get("fos_name") or "").strip()
            matched_emp_id = None
            if raw_fos and raw_fos not in ["-", "Office", "None"]:
                norm_fos = raw_fos.lower()
                if norm_fos in fos_map:
                    matched_emp_id = str(fos_map[norm_fos])
                elif norm_fos in emp_by_name:
                    matched_emp_id = str(emp_by_name[norm_fos].id)
                else:
                    # Fuzzy match first name
                    for e_name, e_obj in emp_by_name.items():
                        if norm_fos in e_name or e_name in norm_fos:
                            matched_emp_id = str(e_obj.id)
                            break
                    if not matched_emp_id:
                        unmatched_fos_names.add(raw_fos)

            brand = str(row_data.get("brand") or "General").strip()
            sales = _parse_decimal(row_data.get("sales_amount"))
            collection = _parse_decimal(row_data.get("collection_amount"))
            market_os = _parse_decimal(row_data.get("market_outstanding"))
            b_lt_15 = _parse_decimal(row_data.get("bucket_lt_15"))
            b_15_30 = _parse_decimal(row_data.get("bucket_15_30"))
            b_30_45 = _parse_decimal(row_data.get("bucket_30_45"))
            b_45_60 = _parse_decimal(row_data.get("bucket_45_60"))
            b_60_75 = _parse_decimal(row_data.get("bucket_60_75"))
            b_75_90 = _parse_decimal(row_data.get("bucket_75_90"))
            b_gt_90 = _parse_decimal(row_data.get("bucket_gt_90"))

            is_existing = dms_code.lower() in cust_by_code
            if is_existing:
                rows_updated += 1
            else:
                rows_created += 1

            plan_rows.append({
                "type": "outlet_bi",
                "dms_code": dms_code,
                "outlet_name": outlet_name or dms_code,
                "zone_name": zone_name,
                "area_name": area_name,
                "raw_fos_name": raw_fos,
                "matched_employee_id": matched_emp_id,
                "brand": brand,
                "sales": str(sales),
                "collection": str(collection),
                "market_outstanding": str(market_os),
                "bucket_lt_15": str(b_lt_15),
                "bucket_15_30": str(b_15_30),
                "bucket_30_45": str(b_30_45),
                "bucket_45_60": str(b_45_60),
                "bucket_60_75": str(b_60_75),
                "bucket_75_90": str(b_75_90),
                "bucket_gt_90": str(b_gt_90),
                "is_existing": is_existing,
            })

    workbook.close()

    batch_summary = {
        "detected_type": detected_type,
        "total_rows": len(plan_rows) + rows_error + rows_skipped,
        "valid_rows": len(plan_rows),
        "unmatched_fos_names": sorted(list(unmatched_fos_names)),
        "plan_rows": plan_rows,
    }

    batch = ImportBatch(
        filename=filename,
        sheet_name=sheet_name,
        uploaded_by=current_user.id,
        column_mapping=column_mapping,
        outlet_match_strategy=outlet_match_strategy,
        status=ImportStatus.VALIDATED,
        summary=batch_summary,
        error_report=error_report,
        rows_processed=len(plan_rows) + rows_error + rows_skipped,
        rows_created=rows_created,
        rows_updated=rows_updated,
        rows_skipped=rows_skipped,
        rows_error=rows_error,
    )
    session.add(batch)
    await session.commit()
    await session.refresh(batch)
    return batch


async def commit_import_batch(
    batch_id: uuid.UUID,
    current_user: User,
    session: AsyncSession,
) -> ImportBatch:
    """
    Executes the validated import plan inside a single transaction.
    """
    batch = await session.get(ImportBatch, batch_id)
    if not batch:
        raise BaseAPIException(status_code=404, detail="Import batch not found", error_code="IMPORT_BATCH_NOT_FOUND")
    if batch.status == ImportStatus.COMMITTED:
        return batch

    summary = batch.summary or {}
    plan_rows = summary.get("plan_rows", [])
    today = date.today()

    try:
        # Load caches
        emp_res = await session.execute(select(Employee))
        emp_by_code = {e.employee_code.lower(): e for e in emp_res.scalars().all() if e.employee_code}

        user_res = await session.execute(select(User))
        user_by_email = {u.email.lower(): u for u in user_res.scalars().all() if u.email}

        terr_res = await session.execute(select(Territory))
        terr_by_name = {t.name.strip().lower(): t for t in terr_res.scalars().all() if t.name}

        area_res = await session.execute(select(Area))
        area_cache = {(a.name.strip().lower(), a.territory_id): a for a in area_res.scalars().all() if a.name}

        cust_res = await session.execute(select(Customer))
        cust_by_code = {c.outlet_code.lower(): c for c in cust_res.scalars().all() if c.outlet_code}

        credentials_sheet_data: list[dict] = []

        for row in plan_rows:
            row_type = row.get("type")

            if row_type == "employee":
                emp_code = row["employee_code"]
                email = row["email"].lower()
                app_role = Role.ADMIN if row["application_role"] == "ADMIN" else Role.EMPLOYEE

                # Find or create User
                user_obj = user_by_email.get(email)
                if not user_obj:
                    user_obj = User(
                        email=email,
                        mobile_number=row.get("mobile_phone") or None,
                        password_hash=hash_password(row["temp_password"]),
                        role=app_role,
                        is_active=True,
                    )
                    session.add(user_obj)
                    await session.flush()
                    user_by_email[email] = user_obj
                else:
                    user_obj.role = app_role

                # Find or create Employee
                emp_obj = emp_by_code.get(emp_code.lower())
                dob_val = date.fromisoformat(row["date_of_birth"]) if row.get("date_of_birth") else None
                if not emp_obj:
                    emp_obj = Employee(
                        user_id=user_obj.id,
                        full_name=row["full_name"],
                        employee_code=emp_code,
                        working_profile=row.get("working_profile"),
                        cug=row.get("cug"),
                        date_of_birth=dob_val,
                        address=row.get("address"),
                        must_change_password=True,
                    )
                    session.add(emp_obj)
                    await session.flush()
                    emp_by_code[emp_code.lower()] = emp_obj
                else:
                    emp_obj.full_name = row["full_name"]
                    emp_obj.working_profile = row.get("working_profile")
                    emp_obj.cug = row.get("cug")
                    emp_obj.date_of_birth = dob_val
                    emp_obj.address = row.get("address")

                credentials_sheet_data.append({
                    "employee_name": row["full_name"],
                    "employee_id": emp_code,
                    "email": email,
                    "temporary_password": row["temp_password"],
                    "application_role": row["application_role"],
                    "working_profile": row.get("working_profile", "FOS"),
                    "cug": row.get("cug", ""),
                })

            elif row_type == "outlet_bi":
                zone_name = row["zone_name"].strip()
                area_name = row["area_name"].strip()
                dms_code = row["dms_code"].strip()
                outlet_name = row["outlet_name"].strip()

                # Zone / Territory
                zone_key = zone_name.lower()
                terr_obj = terr_by_name.get(zone_key)
                if not terr_obj:
                    terr_obj = Territory(name=zone_name, status="ACTIVE")
                    session.add(terr_obj)
                    await session.flush()
                    terr_by_name[zone_key] = terr_obj

                # Area
                area_key = (area_name.lower(), terr_obj.id)
                area_obj = area_cache.get(area_key)
                if not area_obj:
                    area_obj = Area(name=area_name, territory_id=terr_obj.id)
                    session.add(area_obj)
                    await session.flush()
                    area_cache[area_key] = area_obj

                # Customer / Outlet
                dms_key = dms_code.lower()
                cust_obj = cust_by_code.get(dms_key)
                if not cust_obj:
                    cust_obj = Customer(
                        name=outlet_name,
                        outlet_code=dms_code,
                        territory_id=terr_obj.id,
                        area_id=area_obj.id,
                        created_by=current_user.id,
                        location_status="MISSING",
                        contact_number="",
                        address=f"{area_name}, {zone_name}",
                    )
                    session.add(cust_obj)
                    await session.flush()
                    cust_by_code[dms_key] = cust_obj
                else:
                    cust_obj.name = outlet_name
                    cust_obj.territory_id = terr_obj.id
                    cust_obj.area_id = area_obj.id

                # FOS Direct Assignment
                matched_emp_id = row.get("matched_employee_id")
                if matched_emp_id:
                    emp_uuid = uuid.UUID(matched_emp_id)
                    # Check assignment
                    assign_res = await session.execute(
                        select(EmployeeCustomerAssignment).where(
                            EmployeeCustomerAssignment.employee_id == emp_uuid,
                            EmployeeCustomerAssignment.customer_id == cust_obj.id,
                        )
                    )
                    if not assign_res.scalar_one_or_none():
                        session.add(EmployeeCustomerAssignment(
                            employee_id=emp_uuid,
                            customer_id=cust_obj.id,
                            created_by=current_user.id,
                        ))

                # Financial Snapshot
                brand = row.get("brand", "General")
                snap_res = await session.execute(
                    select(OutletFinancialSnapshot).where(
                        OutletFinancialSnapshot.customer_id == cust_obj.id,
                        OutletFinancialSnapshot.brand == brand,
                        OutletFinancialSnapshot.snapshot_date == today,
                    )
                )
                snap_obj = snap_res.scalar_one_or_none()
                if not snap_obj:
                    snap_obj = OutletFinancialSnapshot(
                        customer_id=cust_obj.id,
                        brand=brand,
                        snapshot_date=today,
                        sales=Decimal(row["sales"]),
                        collection=Decimal(row["collection"]),
                        market_outstanding=Decimal(row["market_outstanding"]),
                        bucket_lt_15=Decimal(row["bucket_lt_15"]),
                        bucket_15_30=Decimal(row["bucket_15_30"]),
                        bucket_30_45=Decimal(row["bucket_30_45"]),
                        bucket_45_60=Decimal(row["bucket_45_60"]),
                        bucket_60_75=Decimal(row["bucket_60_75"]),
                        bucket_75_90=Decimal(row["bucket_75_90"]),
                        bucket_gt_90=Decimal(row["bucket_gt_90"]),
                        import_batch_id=batch.id,
                    )
                    session.add(snap_obj)
                else:
                    snap_obj.sales = Decimal(row["sales"])
                    snap_obj.collection = Decimal(row["collection"])
                    snap_obj.market_outstanding = Decimal(row["market_outstanding"])
                    snap_obj.bucket_lt_15 = Decimal(row["bucket_lt_15"])
                    snap_obj.bucket_15_30 = Decimal(row["bucket_15_30"])
                    snap_obj.bucket_30_45 = Decimal(row["bucket_30_45"])
                    snap_obj.bucket_45_60 = Decimal(row["bucket_45_60"])
                    snap_obj.bucket_60_75 = Decimal(row["bucket_60_75"])
                    snap_obj.bucket_75_90 = Decimal(row["bucket_75_90"])
                    snap_obj.bucket_gt_90 = Decimal(row["bucket_gt_90"])
                    snap_obj.import_batch_id = batch.id

        # Update batch summary with generated credentials if employee import
        if credentials_sheet_data:
            summary["credentials"] = credentials_sheet_data

        batch.status = ImportStatus.COMMITTED
        batch.committed_at = datetime.now(timezone.utc)
        batch.committed_by = current_user.id
        batch.summary = summary
        await session.commit()
        await session.refresh(batch)
        return batch

    except Exception as err:
        await session.rollback()
        batch.status = ImportStatus.FAILED
        batch.failure_reason = str(err)
        await session.commit()
        raise BaseAPIException(status_code=500, detail=f"Import commit failed: {err}", error_code="IMPORT_COMMIT_FAILED")


def generate_onboarding_excel(credentials: list[dict]) -> bytes:
    """Creates a stylized Excel spreadsheet for employee onboarding credentials."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Credentials"

    # Header styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    headers = [
        "Employee Name",
        "Employee ID",
        "Email (Login Username)",
        "Temporary Password",
        "Application Role",
        "Working Profile",
        "CUG Number",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cred in credentials:
        ws.append([
            cred.get("employee_name", ""),
            cred.get("employee_id", ""),
            cred.get("email", ""),
            cred.get("temporary_password", ""),
            cred.get("application_role", ""),
            cred.get("working_profile", ""),
            cred.get("cug", ""),
        ])

    # Adjust widths and borders
    for row in ws.iter_rows(min_row=1, max_row=len(credentials) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.row > 1:
                cell.font = data_font
                cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def list_import_batches(session: AsyncSession, skip: int = 0, limit: int = 50) -> list[ImportBatch]:
    res = await session.execute(
        select(ImportBatch).order_by(ImportBatch.uploaded_at.desc()).offset(skip).limit(limit)
    )
    return res.scalars().all()


async def get_import_batch(batch_id: uuid.UUID, session: AsyncSession) -> ImportBatch:
    batch = await session.get(ImportBatch, batch_id)
    if not batch:
        raise BaseAPIException(status_code=404, detail="Import batch not found", error_code="IMPORT_BATCH_NOT_FOUND")
    return batch
