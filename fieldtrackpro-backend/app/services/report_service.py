"""
Report Service: High-performance business intelligence, operational analytics,
and monthly snapshot management.
"""
from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, or_, and_, func, desc, case
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions.custom import BaseAPIException
from app.models.customer import Customer
from app.models.employee import Employee
from app.models.user import User, Role
from app.models.territory import Territory
from app.models.area import Area
from app.models.visit import Visit, VisitStatus
from app.models.geo_verification_log import GeoVerificationLog
from app.models.employee_customer_assignment import EmployeeCustomerAssignment
from app.models.outlet_financial_snapshot import OutletFinancialSnapshot
from app.models.monthly_reporting_period import MonthlyReportingPeriod, MonthlyPeriodStatus
from app.schemas.financial_snapshot import (
    BusinessBIDashboard,
    BusinessSummaryRow,
    MonthlyPeriodRead,
)
from app.schemas.reports import (
    OverviewReportData,
    EmployeeMasterReportRow,
    OutletReportRow,
    OutstandingAgeingReportRow,
    CollectionReportRow,
    VisitDetailedReportRow,
    GeoVerificationReportRow,
)


class ReportService:

    @staticmethod
    async def get_overview_report(
        session: AsyncSession,
        brand: Optional[str] = None,
        zone_id: Optional[uuid.UUID] = None,
        area_id: Optional[uuid.UUID] = None,
        employee_id: Optional[uuid.UUID] = None,
        month: Optional[str] = None,
    ) -> OverviewReportData:
        """
        Consolidated Overview KPIs based on real database records.
        """
        # Get BI Dashboard aggregations for the filter
        bi_dash = await ReportService.get_business_bi_dashboard(
            session=session,
            brand=brand,
            zone_id=zone_id,
            area_id=area_id,
            employee_id=employee_id,
            month=month,
        )

        # Count total active employees
        emp_stmt = select(func.count(Employee.id))
        emp_res = await session.execute(emp_stmt)
        total_employees = emp_res.scalar() or 0

        # Operational visits metrics
        visit_stmt = select(
            func.count(Visit.id).label("total_visits"),
            func.sum(case((Visit.status == VisitStatus.COMPLETED, 1), else_=0)).label("completed_visits"),
        )
        if employee_id:
            visit_stmt = visit_stmt.where(Visit.employee_id == employee_id)
        if zone_id or area_id:
            visit_stmt = visit_stmt.join(Customer, Visit.customer_id == Customer.id)
            if zone_id:
                visit_stmt = visit_stmt.where(Customer.territory_id == zone_id)
            if area_id:
                visit_stmt = visit_stmt.where(Customer.area_id == area_id)

        visit_res = await session.execute(visit_stmt)
        v_row = visit_res.one_or_none()
        total_visits = v_row.total_visits or 0 if v_row else 0
        completed_visits = v_row.completed_visits or 0 if v_row else 0
        completion_rate = round((completed_visits / total_visits * 100), 1) if total_visits > 0 else 0.0

        return OverviewReportData(
            total_employees=total_employees,
            total_outlets=bi_dash.total_outlets,
            total_sales=bi_dash.total_sales,
            total_collection=bi_dash.total_collection,
            total_market_outstanding=bi_dash.total_market_outstanding,
            total_overdue_gt_90=bi_dash.total_overdue_gt_90,
            total_visits=total_visits,
            completed_visits=completed_visits,
            completion_rate=completion_rate,
            brand_breakdown=bi_dash.brand_summaries,
            zone_breakdown=bi_dash.zone_summaries,
            fos_breakdown=bi_dash.fos_summaries,
        )

    @staticmethod
    async def get_employee_master_report(
        session: AsyncSession,
        working_profile: Optional[str] = None,
        role: Optional[str] = None,
        is_active: Optional[bool] = None,
        query: Optional[str] = None,
    ) -> list[EmployeeMasterReportRow]:
        """
        Detailed employee master reporting with real working profile, CUG, and assigned outlet counts.
        """
        stmt = (
            select(
                Employee,
                User.role.label("app_role"),
                User.is_active.label("user_active"),
                User.email.label("user_email"),
                User.mobile_number.label("user_mobile"),
                func.count(EmployeeCustomerAssignment.customer_id).label("assigned_outlets_count"),
            )
            .join(User, Employee.user_id == User.id)
            .outerjoin(EmployeeCustomerAssignment, EmployeeCustomerAssignment.employee_id == Employee.id)
            .group_by(Employee.id, User.id)
            .order_by(Employee.employee_code.asc())
        )

        if working_profile and working_profile != "ALL":
            stmt = stmt.where(func.lower(Employee.working_profile) == func.lower(working_profile))
        if role and role != "ALL":
            stmt = stmt.where(User.role == Role(role))
        if is_active is not None:
            stmt = stmt.where(User.is_active == is_active)
        if query:
            q_str = f"%{query.strip().lower()}%"
            stmt = stmt.where(
                or_(
                    func.lower(Employee.full_name).like(q_str),
                    func.lower(Employee.employee_code).like(q_str),
                    func.lower(User.mobile_number).like(q_str),
                    func.lower(Employee.cug).like(q_str),
                )
            )

        res = await session.execute(stmt)
        rows = res.all()

        out = []
        for emp, app_role, user_act, u_email, u_mobile, outlet_cnt in rows:
            out.append(
                EmployeeMasterReportRow(
                    employee_id=emp.id,
                    employee_code=emp.employee_code,
                    full_name=emp.full_name,
                    email=u_email or "",
                    phone_number=u_mobile or emp.cug or "",
                    cug=emp.cug or "",
                    working_profile=emp.working_profile or "FOS",
                    role=app_role.value if hasattr(app_role, "value") else str(app_role),
                    is_active=user_act,
                    assigned_outlets_count=outlet_cnt or 0,
                    zone_names=[],
                )
            )
        return out

    @staticmethod
    async def get_outlet_report(
        session: AsyncSession,
        brand: Optional[str] = None,
        zone_id: Optional[uuid.UUID] = None,
        area_id: Optional[uuid.UUID] = None,
        employee_id: Optional[uuid.UUID] = None,
        location_status: Optional[str] = None,
        query: Optional[str] = None,
    ) -> list[OutletReportRow]:
        """
        Complete outlet directory report with geographic location and financial aggregates.
        """
        from app.services.customer_service import extract_coords

        stmt = (
            select(
                Customer.id,
                Customer.name,
                Customer.outlet_code,
                Customer.contact_person,
                Customer.contact_number,
                Customer.address,
                Customer.location,
                Customer.geofence_radius_m,
                Customer.location_status,
                Territory.name.label("zone_name"),
                Area.name.label("area_name"),
                Employee.full_name.label("fos_name"),
                func.coalesce(func.sum(OutletFinancialSnapshot.sales), Decimal("0.00")).label("tot_sales"),
                func.coalesce(func.sum(OutletFinancialSnapshot.collection), Decimal("0.00")).label("tot_collection"),
                func.coalesce(func.sum(OutletFinancialSnapshot.market_outstanding), Decimal("0.00")).label("tot_os"),
                func.coalesce(func.sum(OutletFinancialSnapshot.bucket_gt_90), Decimal("0.00")).label("tot_gt_90"),
            )
            .outerjoin(Territory, Customer.territory_id == Territory.id)
            .outerjoin(Area, Customer.area_id == Area.id)
            .outerjoin(EmployeeCustomerAssignment, EmployeeCustomerAssignment.customer_id == Customer.id)
            .outerjoin(Employee, EmployeeCustomerAssignment.employee_id == Employee.id)
            .outerjoin(OutletFinancialSnapshot, OutletFinancialSnapshot.customer_id == Customer.id)
            .group_by(
                Customer.id,
                Customer.name,
                Customer.outlet_code,
                Customer.contact_person,
                Customer.contact_number,
                Customer.address,
                Customer.location,
                Customer.geofence_radius_m,
                Customer.location_status,
                Territory.name,
                Area.name,
                Employee.full_name,
            )
            .order_by(Customer.name.asc())
        )

        if zone_id:
            stmt = stmt.where(Customer.territory_id == zone_id)
        if area_id:
            stmt = stmt.where(Customer.area_id == area_id)
        if employee_id:
            stmt = stmt.where(EmployeeCustomerAssignment.employee_id == employee_id)
        if location_status and location_status != "ALL":
            stmt = stmt.where(Customer.location_status == location_status)
        if query:
            q_str = f"%{query.strip().lower()}%"
            stmt = stmt.where(
                or_(
                    func.lower(Customer.name).like(q_str),
                    func.lower(Customer.outlet_code).like(q_str),
                    func.lower(Customer.address).like(q_str),
                )
            )

        res = await session.execute(stmt)
        rows = res.all()

        out = []
        for (
            c_id, c_name, c_dms, c_contact_person, c_contact_number, c_addr,
            c_loc, c_radius, c_loc_status, z_name, a_name, f_name, s, col, os_amt, gt90
        ) in rows:
            lat, lng = None, None
            if c_loc is not None:
                lat_val, lng_val = extract_coords(c_loc)
                if lat_val != 0.0 or lng_val != 0.0:
                    lat, lng = lat_val, lng_val

            out.append(
                OutletReportRow(
                    customer_id=c_id,
                    dms_code=c_dms or "",
                    outlet_name=c_name,
                    contact_person=c_contact_person,
                    contact_number=c_contact_number,
                    address=c_addr,
                    zone_name=z_name or "Unassigned",
                    area_name=a_name or "Unassigned",
                    fos_name=f_name or "Unassigned",
                    brand=brand or "Multi-Brand",
                    latitude=lat,
                    longitude=lng,
                    geofence_radius_m=c_radius or 75,
                    location_status=c_loc_status or "MISSING",
                    sales=s,
                    collection=col,
                    market_outstanding=os_amt,
                    overdue_gt_90=gt90,
                )
            )
        return out

    @staticmethod
    async def get_outstanding_ageing_report(
        session: AsyncSession,
        brand: Optional[str] = None,
        zone_id: Optional[uuid.UUID] = None,
        area_id: Optional[uuid.UUID] = None,
        employee_id: Optional[uuid.UUID] = None,
        ageing_bucket: Optional[str] = None,
        month: Optional[str] = None,
        query: Optional[str] = None,
    ) -> list[OutstandingAgeingReportRow]:
        """
        Dedicated Market Outstanding & Ageing Buckets report.
        """
        stmt = (
            select(
                OutletFinancialSnapshot,
                Customer.name.label("customer_name"),
                Customer.outlet_code.label("dms_code"),
                Territory.name.label("zone_name"),
                Area.name.label("area_name"),
                Employee.full_name.label("fos_name"),
            )
            .join(Customer, OutletFinancialSnapshot.customer_id == Customer.id)
            .outerjoin(Territory, Customer.territory_id == Territory.id)
            .outerjoin(Area, Customer.area_id == Area.id)
            .outerjoin(EmployeeCustomerAssignment, EmployeeCustomerAssignment.customer_id == Customer.id)
            .outerjoin(Employee, EmployeeCustomerAssignment.employee_id == Employee.id)
            .where(OutletFinancialSnapshot.market_outstanding > 0)
            .order_by(OutletFinancialSnapshot.bucket_gt_90.desc(), OutletFinancialSnapshot.market_outstanding.desc())
        )

        if brand and brand != "ALL":
            stmt = stmt.where(func.lower(OutletFinancialSnapshot.brand) == func.lower(brand))
        if zone_id:
            stmt = stmt.where(Customer.territory_id == zone_id)
        if area_id:
            stmt = stmt.where(Customer.area_id == area_id)
        if employee_id:
            stmt = stmt.where(EmployeeCustomerAssignment.employee_id == employee_id)
        if month:
            # month in YYYY-MM
            try:
                y, m = [int(x) for x in month.split("-")]
                stmt = stmt.where(
                    and_(
                        func.extract("year", OutletFinancialSnapshot.snapshot_date) == y,
                        func.extract("month", OutletFinancialSnapshot.snapshot_date) == m,
                    )
                )
            except Exception:
                pass

        if ageing_bucket and ageing_bucket != "ALL":
            if ageing_bucket in (">90", "gt_90"):
                stmt = stmt.where(OutletFinancialSnapshot.bucket_gt_90 > 0)
            elif ageing_bucket in ("75-90", "75_90"):
                stmt = stmt.where(OutletFinancialSnapshot.bucket_75_90 > 0)
            elif ageing_bucket in ("60-75", "60_75"):
                stmt = stmt.where(OutletFinancialSnapshot.bucket_60_75 > 0)
            elif ageing_bucket in ("45-60", "45_60"):
                stmt = stmt.where(OutletFinancialSnapshot.bucket_45_60 > 0)
            elif ageing_bucket in ("30-45", "30_45"):
                stmt = stmt.where(OutletFinancialSnapshot.bucket_30_45 > 0)
            elif ageing_bucket in ("15-30", "15_30"):
                stmt = stmt.where(OutletFinancialSnapshot.bucket_15_30 > 0)
            elif ageing_bucket in ("<15", "lt_15"):
                stmt = stmt.where(OutletFinancialSnapshot.bucket_lt_15 > 0)

        if query:
            q_str = f"%{query.strip().lower()}%"
            stmt = stmt.where(
                or_(
                    func.lower(Customer.name).like(q_str),
                    func.lower(Customer.outlet_code).like(q_str),
                    func.lower(Territory.name).like(q_str),
                    func.lower(Area.name).like(q_str),
                    func.lower(Employee.full_name).like(q_str),
                )
            )

        res = await session.execute(stmt)
        records = res.all()

        out = []
        seen_snaps = set()
        for snap, c_name, dms, z_name, a_name, f_name in records:
            if not employee_id:
                if snap.id in seen_snaps:
                    continue
                seen_snaps.add(snap.id)

            # Determine highest overdue bucket
            highest = "Normal (<15d)"
            if snap.bucket_gt_90 > 0:
                highest = "Critical (>90d)"
            elif snap.bucket_75_90 > 0:
                highest = "Severe (75-90d)"
            elif snap.bucket_60_75 > 0:
                highest = "High (60-75d)"
            elif snap.bucket_45_60 > 0:
                highest = "Medium (45-60d)"
            elif snap.bucket_30_45 > 0:
                highest = "Low (30-45d)"
            elif snap.bucket_15_30 > 0:
                highest = "Early (15-30d)"

            out.append(
                OutstandingAgeingReportRow(
                    customer_id=snap.customer_id,
                    dms_code=dms or "",
                    outlet_name=c_name,
                    brand=snap.brand,
                    zone_name=z_name or "Unassigned",
                    area_name=a_name or "Unassigned",
                    fos_name=f_name or "Unassigned",
                    market_outstanding=snap.market_outstanding,
                    bucket_lt_15=snap.bucket_lt_15,
                    bucket_15_30=snap.bucket_15_30,
                    bucket_30_45=snap.bucket_30_45,
                    bucket_45_60=snap.bucket_45_60,
                    bucket_60_75=snap.bucket_60_75,
                    bucket_75_90=snap.bucket_75_90,
                    bucket_gt_90=snap.bucket_gt_90,
                    highest_overdue_bucket=highest,
                )
            )
        return out

    @staticmethod
    async def get_collection_report(
        session: AsyncSession,
        brand: Optional[str] = None,
        zone_id: Optional[uuid.UUID] = None,
        area_id: Optional[uuid.UUID] = None,
        employee_id: Optional[uuid.UUID] = None,
        month: Optional[str] = None,
        query: Optional[str] = None,
    ) -> list[CollectionReportRow]:
        """
        Collections report per outlet and brand snapshot.
        """
        stmt = (
            select(
                OutletFinancialSnapshot,
                Customer.name.label("customer_name"),
                Customer.outlet_code.label("dms_code"),
                Territory.name.label("zone_name"),
                Area.name.label("area_name"),
                Employee.full_name.label("fos_name"),
            )
            .join(Customer, OutletFinancialSnapshot.customer_id == Customer.id)
            .outerjoin(Territory, Customer.territory_id == Territory.id)
            .outerjoin(Area, Customer.area_id == Area.id)
            .outerjoin(EmployeeCustomerAssignment, EmployeeCustomerAssignment.customer_id == Customer.id)
            .outerjoin(Employee, EmployeeCustomerAssignment.employee_id == Employee.id)
            .where(OutletFinancialSnapshot.collection > 0)
            .order_by(OutletFinancialSnapshot.collection.desc())
        )

        if brand and brand != "ALL":
            stmt = stmt.where(func.lower(OutletFinancialSnapshot.brand) == func.lower(brand))
        if zone_id:
            stmt = stmt.where(Customer.territory_id == zone_id)
        if area_id:
            stmt = stmt.where(Customer.area_id == area_id)
        if employee_id:
            stmt = stmt.where(EmployeeCustomerAssignment.employee_id == employee_id)
        if month:
            try:
                y, m = [int(x) for x in month.split("-")]
                stmt = stmt.where(
                    and_(
                        func.extract("year", OutletFinancialSnapshot.snapshot_date) == y,
                        func.extract("month", OutletFinancialSnapshot.snapshot_date) == m,
                    )
                )
            except Exception:
                pass
        if query:
            q_str = f"%{query.strip().lower()}%"
            stmt = stmt.where(
                or_(
                    func.lower(Customer.name).like(q_str),
                    func.lower(Customer.outlet_code).like(q_str),
                )
            )

        res = await session.execute(stmt)
        records = res.all()

        out_cols = []
        seen_col_snaps = set()
        for snap, c_name, dms, z_name, a_name, f_name in records:
            if not employee_id:
                if snap.id in seen_col_snaps:
                    continue
                seen_col_snaps.add(snap.id)

            out_cols.append(
                CollectionReportRow(
                    customer_id=snap.customer_id,
                    dms_code=dms or "",
                    outlet_name=c_name,
                    brand=snap.brand,
                    zone_name=z_name or "Unassigned",
                    area_name=a_name or "Unassigned",
                    fos_name=f_name or "Unassigned",
                    collection_amount=snap.collection,
                    sales_amount=snap.sales,
                    snapshot_date=snap.snapshot_date,
                )
            )
        return out_cols

    @staticmethod
    async def get_visits_detailed_report(
        session: AsyncSession,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employee_id: Optional[uuid.UUID] = None,
        zone_id: Optional[uuid.UUID] = None,
        area_id: Optional[uuid.UUID] = None,
        status: Optional[str] = None,
    ) -> list[VisitDetailedReportRow]:
        """
        Detailed operational visits report with GPS verification status.
        """
        stmt = (
            select(
                Visit,
                Employee.full_name.label("employee_name"),
                Customer.name.label("customer_name"),
                Customer.outlet_code.label("dms_code"),
                Territory.name.label("zone_name"),
                Area.name.label("area_name"),
            )
            .join(Employee, Visit.employee_id == Employee.id)
            .join(Customer, Visit.customer_id == Customer.id)
            .outerjoin(Territory, Customer.territory_id == Territory.id)
            .outerjoin(Area, Customer.area_id == Area.id)
            .order_by(Visit.scheduled_at.desc())
        )

        if start_date:
            stmt = stmt.where(Visit.scheduled_at >= datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc))
        if end_date:
            stmt = stmt.where(Visit.scheduled_at <= datetime.combine(end_date, datetime.max.time(), tzinfo=timezone.utc))
        if employee_id:
            stmt = stmt.where(Visit.employee_id == employee_id)
        if zone_id:
            stmt = stmt.where(Customer.territory_id == zone_id)
        if area_id:
            stmt = stmt.where(Customer.area_id == area_id)
        if status and status != "ALL":
            stmt = stmt.where(Visit.status == VisitStatus(status))

        res = await session.execute(stmt)
        rows = res.all()

        out = []
        for v, e_name, c_name, dms, z_name, a_name in rows:
            dur = None
            if v.check_in_at and v.check_out_at:
                dur = max(1, int((v.check_out_at - v.check_in_at).total_seconds() // 60))

            out.append(
                VisitDetailedReportRow(
                    visit_id=v.id,
                    scheduled_at=v.scheduled_at.isoformat() if v.scheduled_at else "",
                    employee_name=e_name,
                    customer_name=c_name,
                    dms_code=dms or "",
                    zone_name=z_name or "Unassigned",
                    area_name=a_name or "Unassigned",
                    status=v.status.value if hasattr(v.status, "value") else str(v.status),
                    check_in_at=v.check_in_at.isoformat() if v.check_in_at else None,
                    check_out_at=v.check_out_at.isoformat() if v.check_out_at else None,
                    duration_minutes=dur,
                    is_gps_verified=bool(v.check_in_location is not None),
                )
            )
        return out

    @staticmethod
    async def get_employee_visit_report(
        session: AsyncSession,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> list[dict]:
        stmt = (
            select(
                Employee.id.label("employee_id"),
                Employee.full_name.label("employee_name"),
                func.count(Visit.id).label("total_visits"),
                func.sum(case((Visit.status == VisitStatus.COMPLETED, 1), else_=0)).label("completed_visits"),
                func.sum(case((Visit.status == VisitStatus.PENDING, 1), else_=0)).label("pending_visits"),
                func.sum(case((Visit.status == VisitStatus.MISSED, 1), else_=0)).label("missed_visits"),
                func.sum(case((Visit.status == VisitStatus.FLAGGED, 1), else_=0)).label("flagged_visits"),
            )
            .join(Visit, Employee.id == Visit.employee_id)
            .group_by(Employee.id, Employee.full_name)
            .order_by(Employee.full_name.asc())
        )

        if start_date:
            stmt = stmt.where(Visit.scheduled_at >= datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc))
        if end_date:
            stmt = stmt.where(Visit.scheduled_at <= datetime.combine(end_date, datetime.max.time(), tzinfo=timezone.utc))

        result = await session.execute(stmt)
        rows = result.all()

        return [
            {
                "employee_id": row.employee_id,
                "employee_name": row.employee_name,
                "total_visits": row.total_visits,
                "completed_visits": row.completed_visits or 0,
                "pending_visits": row.pending_visits or 0,
                "missed_visits": row.missed_visits or 0,
                "flagged_visits": row.flagged_visits or 0,
                "completion_rate": round(
                    ((row.completed_visits or 0) / row.total_visits * 100), 1
                ) if row.total_visits > 0 else 0.0,
            }
            for row in rows
        ]

    @staticmethod
    async def get_customer_visit_history(session: AsyncSession, customer_id: uuid.UUID) -> list[dict]:
        stmt = (
            select(
                Visit.id.label("visit_id"),
                Visit.scheduled_at,
                Visit.status,
                Visit.check_in_at,
                Visit.check_out_at,
                Employee.full_name.label("employee_name"),
            )
            .join(Employee, Visit.employee_id == Employee.id)
            .where(Visit.customer_id == customer_id)
            .order_by(Visit.scheduled_at.desc())
        )
        result = await session.execute(stmt)
        rows = result.all()
        return [
            {
                "visit_id": row.visit_id,
                "scheduled_at": row.scheduled_at.isoformat() if row.scheduled_at else "",
                "status": row.status.value,
                "employee_name": row.employee_name,
                "check_in_at": row.check_in_at.isoformat() if row.check_in_at else None,
                "check_out_at": row.check_out_at.isoformat() if row.check_out_at else None,
            }
            for row in rows
        ]

    @staticmethod
    async def get_productivity_dashboard(session: AsyncSession) -> dict:
        total_emp = (await session.execute(select(func.count(Employee.id)))).scalar() or 0
        active_emp = (
            await session.execute(
                select(func.count(User.id)).where(User.is_active == True, User.role == Role.EMPLOYEE)
            )
        ).scalar() or 0

        today_start = datetime.combine(date.today(), datetime.min.time(), tzinfo=timezone.utc)
        today_end = datetime.combine(date.today(), datetime.max.time(), tzinfo=timezone.utc)

        stmt = (
            select(
                func.count(Visit.id).label("total"),
                func.sum(case((Visit.status == VisitStatus.COMPLETED, 1), else_=0)).label("completed"),
                func.sum(case((Visit.status == VisitStatus.PENDING, 1), else_=0)).label("pending"),
                func.sum(case((Visit.status == VisitStatus.MISSED, 1), else_=0)).label("missed"),
                func.sum(case((Visit.status == VisitStatus.FLAGGED, 1), else_=0)).label("flagged"),
            )
            .where(Visit.scheduled_at >= today_start, Visit.scheduled_at <= today_end)
        )
        result = await session.execute(stmt)
        row = result.one_or_none()

        tot = row.total if row and row.total else 0
        comp = row.completed if row and row.completed else 0
        pend = row.pending if row and row.pending else 0
        miss = row.missed if row and row.missed else 0
        flag = row.flagged if row and row.flagged else 0

        return {
            "total_employees": total_emp,
            "active_employees": active_emp,
            "total_visits_today": tot,
            "completed_visits_today": comp,
            "pending_visits_today": pend,
            "missed_visits_today": miss,
            "flagged_visits_today": flag,
            "avg_visits_per_employee": round(tot / active_emp, 1) if active_emp > 0 else 0.0,
        }

    @staticmethod
    async def get_geo_verification_report(
        session: AsyncSession,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> list[dict]:
        query = (
            select(
                GeoVerificationLog,
                Employee.full_name.label("employee_name"),
                Customer.name.label("customer_name"),
                Customer.outlet_code.label("dms_code"),
            )
            .join(Visit, GeoVerificationLog.visit_id == Visit.id)
            .join(Employee, Visit.employee_id == Employee.id)
            .join(Customer, Visit.customer_id == Customer.id)
            .order_by(GeoVerificationLog.attempted_at.desc())
        )

        if start_date:
            query = query.where(GeoVerificationLog.attempted_at >= datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc))
        if end_date:
            query = query.where(GeoVerificationLog.attempted_at < datetime.combine(end_date, datetime.max.time(), tzinfo=timezone.utc))

        result = await session.execute(query)
        rows = result.all()

        return [
            {
                "visit_id": str(row.GeoVerificationLog.visit_id) if row.GeoVerificationLog.visit_id else "",
                "employee_name": row.employee_name,
                "customer_name": row.customer_name,
                "dms_code": row.dms_code or "",
                "attempted_at": row.GeoVerificationLog.attempted_at.isoformat(),
                "verification_type": row.GeoVerificationLog.verification_type.value,
                "is_valid": row.GeoVerificationLog.is_valid,
                "distance_m": round(float(row.GeoVerificationLog.distance_from_customer_m), 1) if row.GeoVerificationLog.distance_from_customer_m is not None else 0.0,
                "failure_reason": row.GeoVerificationLog.failure_reason,
            }
            for row in rows
        ]

    @staticmethod
    async def get_business_bi_dashboard(
        session: AsyncSession,
        brand: Optional[str] = None,
        zone_id: Optional[uuid.UUID] = None,
        area_id: Optional[uuid.UUID] = None,
        employee_id: Optional[uuid.UUID] = None,
        month: Optional[str] = None,
    ) -> BusinessBIDashboard:
        """
        Calculates live or historical aggregated Business BI summaries.
        """
        stmt = (
            select(
                OutletFinancialSnapshot,
                Customer.name.label("customer_name"),
                Customer.outlet_code.label("dms_code"),
                Territory.name.label("zone_name"),
                Area.name.label("area_name"),
                Employee.full_name.label("fos_name"),
            )
            .join(Customer, OutletFinancialSnapshot.customer_id == Customer.id)
            .outerjoin(Territory, Customer.territory_id == Territory.id)
            .outerjoin(Area, Customer.area_id == Area.id)
            .outerjoin(EmployeeCustomerAssignment, EmployeeCustomerAssignment.customer_id == Customer.id)
            .outerjoin(Employee, EmployeeCustomerAssignment.employee_id == Employee.id)
        )

        if brand and brand != "ALL":
            stmt = stmt.where(func.lower(OutletFinancialSnapshot.brand) == func.lower(brand))
        if zone_id:
            stmt = stmt.where(Customer.territory_id == zone_id)
        if area_id:
            stmt = stmt.where(Customer.area_id == area_id)
        if employee_id:
            stmt = stmt.where(EmployeeCustomerAssignment.employee_id == employee_id)
        if month:
            try:
                y, m = [int(x) for x in month.split("-")]
                stmt = stmt.where(
                    and_(
                        func.extract("year", OutletFinancialSnapshot.snapshot_date) == y,
                        func.extract("month", OutletFinancialSnapshot.snapshot_date) == m,
                    )
                )
            except Exception:
                pass

        res = await session.execute(stmt)
        records = res.all()

        total_sales = Decimal("0.00")
        total_collection = Decimal("0.00")
        total_market_os = Decimal("0.00")
        total_gt_90 = Decimal("0.00")
        unique_outlets: set[uuid.UUID] = set()

        raw_rows: list[BusinessSummaryRow] = []
        brand_map: dict[str, dict] = {}
        zone_map: dict[str, dict] = {}
        area_map: dict[str, dict] = {}
        fos_map: dict[str, dict] = {}

        def _init_agg(name: str, b_name: str) -> dict:
            return {
                "dimension_name": name,
                "brand": b_name,
                "outlets": set(),
                "sales": Decimal("0.00"),
                "collection": Decimal("0.00"),
                "market_os": Decimal("0.00"),
                "b_lt_15": Decimal("0.00"),
                "b_15_30": Decimal("0.00"),
                "b_30_45": Decimal("0.00"),
                "b_45_60": Decimal("0.00"),
                "b_60_75": Decimal("0.00"),
                "b_75_90": Decimal("0.00"),
                "b_gt_90": Decimal("0.00"),
            }

        def _add_to_agg(agg: dict, snap: OutletFinancialSnapshot, cust_id: uuid.UUID):
            agg["outlets"].add(cust_id)
            agg["sales"] += snap.sales
            agg["collection"] += snap.collection
            agg["market_os"] += snap.market_outstanding
            agg["b_lt_15"] += snap.bucket_lt_15
            agg["b_15_30"] += snap.bucket_15_30
            agg["b_30_45"] += snap.bucket_30_45
            agg["b_45_60"] += snap.bucket_45_60
            agg["b_60_75"] += snap.bucket_60_75
            agg["b_75_90"] += snap.bucket_75_90
            agg["b_gt_90"] += snap.bucket_gt_90

        seen_snapshots: set[uuid.UUID] = set()
        seen_fos_entries: set[tuple[uuid.UUID, str]] = set()

        for snap, c_name, dms, z_name, a_name, f_name in records:
            b_name = snap.brand or "General"
            z_label = z_name or "Unknown Zone"
            a_label = a_name or "Unknown Area"
            f_label = f_name or "Unassigned"
            snap_id = snap.id

            if snap_id not in seen_snapshots:
                seen_snapshots.add(snap_id)
                unique_outlets.add(snap.customer_id)
                total_sales += snap.sales
                total_collection += snap.collection
                total_market_os += snap.market_outstanding
                total_gt_90 += snap.bucket_gt_90

                raw_rows.append(
                    BusinessSummaryRow(
                        brand=b_name,
                        dimension_name=c_name,
                        dms_code=dms or "",
                        outlet_name=c_name,
                        zone_name=z_label,
                        area_name=a_label,
                        fos_name=f_label,
                        outlets_count=1,
                        sales=snap.sales,
                        collection=snap.collection,
                        market_outstanding=snap.market_outstanding,
                        bucket_lt_15=snap.bucket_lt_15,
                        bucket_15_30=snap.bucket_15_30,
                        bucket_30_45=snap.bucket_30_45,
                        bucket_45_60=snap.bucket_45_60,
                        bucket_60_75=snap.bucket_60_75,
                        bucket_75_90=snap.bucket_75_90,
                        bucket_gt_90=snap.bucket_gt_90,
                    )
                )

                # Brand aggregation (exactly once per snapshot)
                if b_name not in brand_map:
                    brand_map[b_name] = _init_agg(b_name, b_name)
                _add_to_agg(brand_map[b_name], snap, snap.customer_id)

                # Zone aggregation (exactly once per snapshot)
                z_key = f"{b_name}::{z_label}"
                if z_key not in zone_map:
                    zone_map[z_key] = _init_agg(z_label, b_name)
                _add_to_agg(zone_map[z_key], snap, snap.customer_id)

                # Area aggregation (exactly once per snapshot)
                a_key = f"{b_name}::{a_label}"
                if a_key not in area_map:
                    area_map[a_key] = _init_agg(a_label, b_name)
                _add_to_agg(area_map[a_key], snap, snap.customer_id)

            # FOS aggregation (exactly once per (snapshot, FOS) assignment)
            fos_entry_key = (snap_id, f_label)
            if fos_entry_key not in seen_fos_entries:
                seen_fos_entries.add(fos_entry_key)
                f_key = f"{b_name}::{f_label}"
                if f_key not in fos_map:
                    fos_map[f_key] = _init_agg(f_label, b_name)
                _add_to_agg(fos_map[f_key], snap, snap.customer_id)

        def _to_rows(d_map: dict) -> list[BusinessSummaryRow]:
            out = []
            for item in d_map.values():
                out.append(
                    BusinessSummaryRow(
                        brand=item["brand"],
                        dimension_name=item["dimension_name"],
                        outlets_count=len(item["outlets"]),
                        sales=item["sales"],
                        collection=item["collection"],
                        market_outstanding=item["market_os"],
                        bucket_lt_15=item["b_lt_15"],
                        bucket_15_30=item["b_15_30"],
                        bucket_30_45=item["b_30_45"],
                        bucket_45_60=item["b_45_60"],
                        bucket_60_75=item["b_60_75"],
                        bucket_75_90=item["b_75_90"],
                        bucket_gt_90=item["b_gt_90"],
                    )
                )
            return sorted(out, key=lambda x: x.sales, reverse=True)

        # Check finalization status if month specified
        is_fin = False
        if month:
            try:
                y, m = [int(x) for x in month.split("-")]
                p_stmt = select(MonthlyReportingPeriod).where(
                    MonthlyReportingPeriod.period_year == y,
                    MonthlyReportingPeriod.period_month == m,
                )
                p_res = await session.execute(p_stmt)
                period_row = p_res.scalar_one_or_none()
                if period_row and period_row.status == MonthlyPeriodStatus.FINALIZED:
                    is_finalized = True
            except Exception:
                pass

        return BusinessBIDashboard(
            snapshot_date=date.today(),
            month_period=month,
            is_finalized=is_fin,
            total_outlets=len(unique_outlets),
            total_sales=total_sales,
            total_collection=total_collection,
            total_market_outstanding=total_market_os,
            total_overdue_gt_90=total_gt_90,
            brand_summaries=_to_rows(brand_map),
            zone_summaries=_to_rows(zone_map),
            area_summaries=_to_rows(area_map),
            fos_summaries=_to_rows(fos_map),
            raw_outlet_rows=raw_rows,
        )

    # -----------------------------------------------------------------------
    # PHASE 4: Monthly Reporting Periods & Historical Snapshots
    # -----------------------------------------------------------------------
    @staticmethod
    async def get_monthly_periods(session: AsyncSession) -> list[MonthlyPeriodRead]:
        """
        Discovers snapshot months from database, computes and syncs monthly period records.
        """
        # Find distinct year-month pairs from outlet_financial_snapshots
        snap_stmt = select(
            func.extract("year", OutletFinancialSnapshot.snapshot_date).label("s_year"),
            func.extract("month", OutletFinancialSnapshot.snapshot_date).label("s_month"),
            func.count(OutletFinancialSnapshot.id).label("snap_cnt"),
            func.count(func.distinct(OutletFinancialSnapshot.customer_id)).label("outlets_cnt"),
            func.sum(OutletFinancialSnapshot.sales).label("s_sales"),
            func.sum(OutletFinancialSnapshot.collection).label("s_collection"),
            func.sum(OutletFinancialSnapshot.market_outstanding).label("s_os"),
            func.sum(OutletFinancialSnapshot.bucket_gt_90).label("s_gt90"),
        ).group_by(
            func.extract("year", OutletFinancialSnapshot.snapshot_date),
            func.extract("month", OutletFinancialSnapshot.snapshot_date),
        ).order_by(
            func.extract("year", OutletFinancialSnapshot.snapshot_date).desc(),
            func.extract("month", OutletFinancialSnapshot.snapshot_date).desc(),
        )

        res = await session.execute(snap_stmt)
        snap_months = res.all()

        month_names = [
            "", "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]

        # Sync or ensure MonthlyReportingPeriod exists for each month
        for s_year, s_month, snap_cnt, out_cnt, tot_s, tot_c, tot_os, tot_gt90 in snap_months:
            y = int(s_year)
            m = int(s_month)
            p_name = f"{month_names[m]} {y}"

            p_stmt = select(MonthlyReportingPeriod).where(
                MonthlyReportingPeriod.period_year == y,
                MonthlyReportingPeriod.period_month == m,
            )
            p_res = await session.execute(p_stmt)
            period = p_res.scalar_one_or_none()

            if not period:
                period = MonthlyReportingPeriod(
                    period_year=y,
                    period_month=m,
                    period_name=p_name,
                    status=MonthlyPeriodStatus.OPEN,
                    snapshot_count=snap_cnt or 0,
                    total_outlets=out_cnt or 0,
                    total_sales=tot_s or Decimal("0.00"),
                    total_collection=tot_c or Decimal("0.00"),
                    total_market_os=tot_os or Decimal("0.00"),
                    total_overdue_gt_90=tot_gt90 or Decimal("0.00"),
                )
                session.add(period)
            else:
                # Update counts only if OPEN (preserve finalized numbers if locked)
                if period.status == MonthlyPeriodStatus.OPEN:
                    period.snapshot_count = snap_cnt or 0
                    period.total_outlets = out_cnt or 0
                    period.total_sales = tot_s or Decimal("0.00")
                    period.total_collection = tot_c or Decimal("0.00")
                    period.total_market_os = tot_os or Decimal("0.00")
                    period.total_overdue_gt_90 = tot_gt90 or Decimal("0.00")

        await session.flush()

        # Fetch all periods
        all_periods = (
            await session.execute(
                select(MonthlyReportingPeriod).order_by(
                    MonthlyReportingPeriod.period_year.desc(),
                    MonthlyReportingPeriod.period_month.desc(),
                )
            )
        ).scalars().all()

        return [
            MonthlyPeriodRead(
                id=p.id,
                period_year=p.period_year,
                period_month=p.period_month,
                period_name=p.period_name,
                status=p.status.value if hasattr(p.status, "value") else str(p.status),
                snapshot_count=p.snapshot_count,
                total_outlets=p.total_outlets,
                total_sales=p.total_sales,
                total_collection=p.total_collection,
                total_market_os=p.total_market_os,
                total_overdue_gt_90=p.total_overdue_gt_90,
                finalized_at=p.finalized_at,
                finalized_by=p.finalized_by,
                created_at=p.created_at,
                updated_at=p.updated_at,
            )
            for p in all_periods
        ]

    @staticmethod
    async def finalize_monthly_period(
        session: AsyncSession,
        period_id: uuid.UUID,
        user_id: uuid.UUID,
    ) -> MonthlyPeriodRead:
        """
        Finalizes a monthly reporting period. Locks historical data.
        """
        stmt = select(MonthlyReportingPeriod).where(MonthlyReportingPeriod.id == period_id)
        res = await session.execute(stmt)
        period = res.scalar_one_or_none()
        if not period:
            raise BaseAPIException(status_code=404, detail="Monthly period not found", error_code="PERIOD_NOT_FOUND")

        now_dt = datetime.now(timezone.utc)
        period.status = MonthlyPeriodStatus.FINALIZED
        period.finalized_at = now_dt
        period.finalized_by = user_id
        period.updated_at = now_dt
        await session.flush()

        return MonthlyPeriodRead(
            id=period.id,
            period_year=period.period_year,
            period_month=period.period_month,
            period_name=period.period_name,
            status="FINALIZED",
            snapshot_count=period.snapshot_count,
            total_outlets=period.total_outlets,
            total_sales=period.total_sales,
            total_collection=period.total_collection,
            total_market_os=period.total_market_os,
            total_overdue_gt_90=period.total_overdue_gt_90,
            finalized_at=now_dt,
            finalized_by=user_id,
            created_at=period.created_at or now_dt,
            updated_at=now_dt,
        )

    @staticmethod
    async def reopen_monthly_period(
        session: AsyncSession,
        period_id: uuid.UUID,
    ) -> MonthlyPeriodRead:
        """
        Reopens a finalized monthly reporting period.
        """
        stmt = select(MonthlyReportingPeriod).where(MonthlyReportingPeriod.id == period_id)
        res = await session.execute(stmt)
        period = res.scalar_one_or_none()
        if not period:
            raise BaseAPIException(status_code=404, detail="Monthly period not found", error_code="PERIOD_NOT_FOUND")

        now_dt = datetime.now(timezone.utc)
        period.status = MonthlyPeriodStatus.OPEN
        period.finalized_at = None
        period.finalized_by = None
        period.updated_at = now_dt
        await session.flush()

        return MonthlyPeriodRead(
            id=period.id,
            period_year=period.period_year,
            period_month=period.period_month,
            period_name=period.period_name,
            status="OPEN",
            snapshot_count=period.snapshot_count,
            total_outlets=period.total_outlets,
            total_sales=period.total_sales,
            total_collection=period.total_collection,
            total_market_os=period.total_market_os,
            total_overdue_gt_90=period.total_overdue_gt_90,
            finalized_at=None,
            finalized_by=None,
            created_at=period.created_at or now_dt,
            updated_at=now_dt,
        )

    # -----------------------------------------------------------------------
    # Excel Export Engine (Business-friendly formatting)
    # -----------------------------------------------------------------------
    @staticmethod
    def export_report_excel(
        report_title: str,
        sheet_name: str,
        headers: list[str],
        data_rows: list[list[Any]],
        filters_applied: dict[str, str] | None = None,
        summary_kpis: list[tuple[str, str]] | None = None,
    ) -> bytes:
        """
        Generates business-ready, styled Excel workbook for any report type.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        # Visual design tokens
        title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        sub_font = Font(name="Calibri", size=9, italic=True, color="6B7280")
        kpi_lbl_font = Font(name="Calibri", size=9, bold=True, color="4B5563")
        kpi_val_font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
        hdr_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        total_font = Font(name="Calibri", size=10, bold=True, color="111827")
        data_font = Font(name="Calibri", size=10)

        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        # 1. Report Title & Header Banner
        ws.append([f"FieldTrack — {report_title}"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.cell(row=2, column=1).font = sub_font

        # 2. Applied Filters block
        if filters_applied:
            f_str = " | ".join(f"{k}: {v}" for k, v in filters_applied.items() if v and v != "ALL")
            if f_str:
                ws.append([f"Filters: {f_str}"])
                ws.cell(row=3, column=1).font = sub_font

        ws.append([])

        # 3. Summary KPIs Banner
        if summary_kpis:
            kpi_labels = [k for k, _ in summary_kpis]
            kpi_values = [v for _, v in summary_kpis]
            ws.append(kpi_labels)
            lbl_row = ws.max_row
            ws.append(kpi_values)
            val_row = ws.max_row

            for col_i in range(1, len(summary_kpis) + 1):
                c_lbl = ws.cell(row=lbl_row, column=col_i)
                c_lbl.font = kpi_lbl_font
                c_lbl.fill = total_fill
                c_val = ws.cell(row=val_row, column=col_i)
                c_val.font = kpi_val_font
                c_val.fill = total_fill

            ws.append([])

        # 4. Data Table Headers
        ws.append(headers)
        header_row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=header_row_idx, column=col_idx)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        # 5. Data Rows
        for r in data_rows:
            ws.append(r)
            cur_r = ws.max_row
            for col_idx in range(1, len(r) + 1):
                cell = ws.cell(row=cur_r, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                val = cell.value
                if isinstance(val, (int, float, Decimal)):
                    cell.number_format = "#,##0.00" if isinstance(val, (float, Decimal)) else "#,##0"
                    cell.alignment = Alignment(horizontal="right")

        # Auto column width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or "")
                if cell.row > 4 and len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def export_business_bi_excel(dashboard: BusinessBIDashboard) -> bytes:
        """
        Exports the multi-sheet Business BI workbook.
        """
        wb = openpyxl.Workbook()

        hdr_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
        sub_font = Font(name="Calibri", size=9, italic=True)
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        def _write_section(ws, title: str, headers: list[str], rows: list[BusinessSummaryRow], is_raw: bool = False):
            ws.append([title])
            ws.cell(row=ws.max_row, column=1).font = title_font
            ws.append([f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Outlets: {dashboard.total_outlets} | Total Sales: ₹{dashboard.total_sales:,.2f} | Total OS: ₹{dashboard.total_market_outstanding:,.2f}"])
            ws.cell(row=ws.max_row, column=1).font = sub_font
            ws.append([])
            ws.append(headers)
            header_row_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=header_row_idx, column=col_idx)
                c.fill = hdr_fill
                c.font = hdr_font
                c.alignment = Alignment(horizontal="center", vertical="center")

            for r in rows:
                if is_raw:
                    row_data = [
                        r.brand, r.dms_code, r.outlet_name, r.zone_name, r.area_name, r.fos_name,
                        float(r.sales), float(r.collection), float(r.market_outstanding),
                        float(r.bucket_lt_15), float(r.bucket_15_30), float(r.bucket_30_45),
                        float(r.bucket_45_60), float(r.bucket_60_75), float(r.bucket_75_90),
                        float(r.bucket_gt_90),
                    ]
                else:
                    row_data = [
                        r.brand, r.dimension_name, r.outlets_count,
                        float(r.sales), float(r.collection), float(r.market_outstanding),
                        float(r.bucket_lt_15), float(r.bucket_15_30), float(r.bucket_30_45),
                        float(r.bucket_45_60), float(r.bucket_60_75), float(r.bucket_75_90),
                        float(r.bucket_gt_90),
                    ]
                ws.append(row_data)
                cur_r = ws.max_row
                for c_i in range(1, len(row_data) + 1):
                    cell = ws.cell(row=cur_r, column=c_i)
                    cell.border = thin_border
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.00"

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col if cell.row > 3) if col else 10
                ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

        # 1. Brand Summary Sheet
        ws_brand = wb.active
        ws_brand.title = "Brand Summary"
        _write_section(
            ws_brand,
            "Brand-Wise Performance Summary",
            ["Brand", "Dimension", "Outlets", "Sales", "Collection", "Market OS", "<15d", "15-30d", "30-45d", "45-60d", "60-75d", "75-90d", ">90d"],
            dashboard.brand_summaries,
        )

        # 2. Zone Summary Sheet
        ws_zone = wb.create_sheet("Zone Summary")
        _write_section(
            ws_zone,
            "Zone-Wise Performance Summary",
            ["Brand", "Zone Name", "Outlets", "Sales", "Collection", "Market OS", "<15d", "15-30d", "30-45d", "45-60d", "60-75d", "75-90d", ">90d"],
            dashboard.zone_summaries,
        )

        # 3. Area Summary Sheet
        ws_area = wb.create_sheet("Area Summary")
        _write_section(
            ws_area,
            "Area-Wise Performance Summary",
            ["Brand", "Area Name", "Outlets", "Sales", "Collection", "Market OS", "<15d", "15-30d", "30-45d", "45-60d", "60-75d", "75-90d", ">90d"],
            dashboard.area_summaries,
        )

        # 4. FOS Summary Sheet
        ws_fos = wb.create_sheet("FOS Summary")
        _write_section(
            ws_fos,
            "FOS / Field Officer Performance Summary",
            ["Brand", "FOS Name", "Outlets", "Sales", "Collection", "Market OS", "<15d", "15-30d", "30-45d", "45-60d", "60-75d", "75-90d", ">90d"],
            dashboard.fos_summaries,
        )

        # 5. Raw Outlet Data Sheet
        ws_raw = wb.create_sheet("Outlet Data")
        _write_section(
            ws_raw,
            "Complete Outlet Level Financial Master",
            ["Brand", "DMS Code", "Outlet Name", "Zone", "Area", "FOS", "Sales", "Collection", "Market OS", "<15d", "15-30d", "30-45d", "45-60d", "60-75d", "75-90d", ">90d"],
            dashboard.raw_outlet_rows,
            is_raw=True,
        )

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


report_service = ReportService()
