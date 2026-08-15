"""
Integration tests for the Excel/MIS import pipeline: parsing, mapping,
validation, transactional commit, idempotency, and authorization.
"""
from __future__ import annotations

import io
import uuid

import openpyxl
import pytest
import pytest_asyncio
from httpx import AsyncClient

from tests.integration.conftest import db_cursor, requires_db

pytestmark = [requires_db, pytest.mark.integration, pytest.mark.asyncio]


def _xlsx(headers: list[str], rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


FULL_HEADERS = [
    "Outlet Code", "Outlet Name", "Address", "Contact", "Latitude", "Longitude",
    "Emp Code", "Territory", "Invoice No", "Invoice Date", "Amount", "Brand",
    "Payment Amount", "Payment Date", "Payment Mode", "Reference",
]

BASE_MAPPING = {
    "Outlet Code": "outlet_code", "Outlet Name": "outlet_name", "Address": "outlet_address",
    "Contact": "outlet_contact_number", "Latitude": "outlet_latitude", "Longitude": "outlet_longitude",
    "Emp Code": "employee_code", "Territory": "territory_name", "Invoice No": "invoice_number",
    "Invoice Date": "invoice_date", "Amount": "invoice_amount", "Brand": "invoice_brand",
    "Payment Amount": "payment_amount", "Payment Date": "payment_date",
    "Payment Mode": "payment_method", "Reference": "payment_reference",
}


@pytest_asyncio.fixture
async def import_cleanup():
    """Track created outlet_codes/territory names/import filenames for teardown."""
    state = {"outlet_codes": [], "territory_names": [], "filenames": []}
    yield state
    with db_cursor(privileged=True) as cur:
        for code in state["outlet_codes"]:
            cur.execute(
                "DELETE FROM payments WHERE customer_id IN (SELECT id FROM customers WHERE outlet_code = %s)", (code,)
            )
            cur.execute(
                "DELETE FROM invoices WHERE customer_id IN (SELECT id FROM customers WHERE outlet_code = %s)", (code,)
            )
            cur.execute("DELETE FROM customers WHERE outlet_code = %s", (code,))
        for name in state["territory_names"]:
            cur.execute("DELETE FROM territories WHERE name = %s", (name,))
        for filename in state["filenames"]:
            cur.execute("DELETE FROM import_batches WHERE filename = %s", (filename,))


def _unique(tag: str) -> str:
    return f"{tag}-{uuid.uuid4().hex[:8]}"


async def _validate(client, admin_headers, file_bytes, filename, mapping, sheet_name="Sheet1", strategy="outlet_code", allow_generated=False):
    import json as _json
    request = _json.dumps({
        "sheet_name": sheet_name, "column_mapping": mapping,
        "outlet_match_strategy": strategy, "allow_generated_invoice_numbers": allow_generated,
    })
    return await client.post(
        "/api/v1/imports/validate",
        files={"file": (filename, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"request": request},
        headers=admin_headers,
    )


# -- Parsing --------------------------------------------------------------------

async def test_preview_lists_sheets_and_suggests_mapping(client: AsyncClient, admin_headers):
    data = _xlsx(FULL_HEADERS, [["A", "B"]], sheet_name="MyData")
    resp = await client.post(
        "/api/v1/imports/preview",
        files={"file": ("f.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["all_sheets"] == ["MyData"]
    assert body["suggested_mapping"]["Outlet Code"] == "outlet_code"
    assert body["suggested_mapping"]["Invoice No"] == "invoice_number"


async def test_preview_rejects_empty_sheet(client: AsyncClient, admin_headers):
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    resp = await client.post(
        "/api/v1/imports/preview",
        files={"file": ("empty.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers,
    )
    assert resp.status_code == 400


async def test_validate_rejects_malformed_mapping_reference(client: AsyncClient, admin_headers):
    data = _xlsx(FULL_HEADERS, [["A"] * len(FULL_HEADERS)])
    resp = await _validate(client, admin_headers, data, "f.xlsx", {"Nonexistent Column": "outlet_code"})
    assert resp.status_code == 400


async def test_validate_handles_multiple_sheets(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("SHEET")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("multi.xlsx")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Other"
    ws1.append(["Junk"])
    ws2 = wb.create_sheet("RealData")
    ws2.append(FULL_HEADERS)
    ws2.append([code, "Sheet Test Outlet", "Addr", "+919800000001", "12.0", "77.0", "", "", "INV-S1", "2026-06-01", "1000", "Usha", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)

    resp = await _validate(client, admin_headers, buf.getvalue(), "multi.xlsx", BASE_MAPPING, sheet_name="RealData")
    assert resp.status_code == 200, resp.text
    assert resp.json()["rows_processed"] == 1


# -- Mapping / validation ---------------------------------------------------------

async def test_missing_outlet_code_becomes_row_error(client: AsyncClient, admin_headers, import_cleanup):
    import_cleanup["filenames"].append("nocode.xlsx")
    data = _xlsx(FULL_HEADERS, [["", "No Code Outlet", "Addr", "+919800000002", "12.0", "77.0", "", "", "INV-X1", "2026-06-01", "500", "Usha", "", "", "", ""]])
    resp = await _validate(client, admin_headers, data, "nocode.xlsx", BASE_MAPPING)
    body = resp.json()
    assert body["rows_error"] == 1
    assert any("outlet_code" in e["error"] for e in body["error_report"])


async def test_invalid_date_is_a_row_error_not_a_crash(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("BADDATE")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("baddate.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "Bad Date Outlet", "Addr", "+919800000003", "12.0", "77.0", "", "", "INV-X2", "not-a-date", "500", "Usha", "", "", "", ""]])
    resp = await _validate(client, admin_headers, data, "baddate.xlsx", BASE_MAPPING)
    assert resp.status_code == 200
    body = resp.json()
    assert body["rows_error"] == 1
    assert any("invoice_date" in e["error"] for e in body["error_report"])


async def test_invalid_amount_is_a_row_error(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("BADAMT")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("badamt.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "Bad Amount Outlet", "Addr", "+919800000004", "12.0", "77.0", "", "", "INV-X3", "2026-06-01", "not-a-number", "Usha", "", "", "", ""]])
    resp = await _validate(client, admin_headers, data, "badamt.xlsx", BASE_MAPPING)
    body = resp.json()
    assert body["rows_error"] == 1
    assert any("invoice_amount" in e["error"] for e in body["error_report"])


async def test_unknown_employee_code_is_a_warning_not_a_blocking_error_without_payment(
    client: AsyncClient, admin_headers, import_cleanup
):
    """An unresolved employee only blocks the row if a payment actually needs it."""
    code = _unique("NOEMP")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("noemp.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "No Employee Outlet", "Addr", "+919800000005", "12.0", "77.0", "GHOST-999", "", "INV-X4", "2026-06-01", "500", "Usha", "", "", "", ""]])
    resp = await _validate(client, admin_headers, data, "noemp.xlsx", BASE_MAPPING)
    body = resp.json()
    assert body["rows_error"] == 0
    assert body["summary"]["invoices_created"] == 1


async def test_unknown_employee_blocks_row_when_payment_present(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("NOEMPPAY")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("noemppay.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "No Employee Payment Outlet", "Addr", "+919800000006", "12.0", "77.0", "GHOST-999", "", "INV-X5", "2026-06-01", "500", "Usha", "200", "2026-06-02", "CASH", ""]])
    resp = await _validate(client, admin_headers, data, "noemppay.xlsx", BASE_MAPPING)
    body = resp.json()
    assert body["rows_error"] == 1
    assert any("payment" in e["error"].lower() for e in body["error_report"])


async def test_unknown_territory_is_auto_created_not_an_error(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("NEWTERR")
    territory_name = _unique("NewTerritory")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["territory_names"].append(territory_name)
    import_cleanup["filenames"].append("newterr.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "New Territory Outlet", "Addr", "+919800000007", "12.0", "77.0", "", territory_name, "INV-X6", "2026-06-01", "500", "Usha", "", "", "", ""]])
    resp = await _validate(client, admin_headers, data, "newterr.xlsx", BASE_MAPPING)
    body = resp.json()
    assert body["rows_error"] == 0
    assert body["summary"]["territories_created"] == 1


async def test_duplicate_invoice_within_same_file_is_flagged_and_not_double_planned(
    client: AsyncClient, admin_headers, import_cleanup
):
    code = _unique("DUPINV")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("dupinv.xlsx")
    row = [code, "Dup Invoice Outlet", "Addr", "+919800000008", "12.0", "77.0", "", "", "INV-DUP1", "2026-06-01", "500", "Usha", "", "", "", ""]
    data = _xlsx(FULL_HEADERS, [row, row])
    resp = await _validate(client, admin_headers, data, "dupinv.xlsx", BASE_MAPPING)
    body = resp.json()
    assert body["rows_error"] == 0
    assert body["summary"]["invoices_created"] == 1
    assert body["summary"]["invoices_skipped_duplicate"] == 1


async def test_missing_outlet_creation_data_is_a_warning_not_a_fabricated_location(
    client: AsyncClient, admin_headers, import_cleanup
):
    """No lat/lng for a brand-new outlet must not silently create a fake location."""
    code = _unique("NOLOC")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("noloc.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "No Location Outlet", "Addr", "+919800000009", "", "", "", "", "INV-X7", "2026-06-01", "500", "Usha", "", "", "", ""]])
    resp = await _validate(client, admin_headers, data, "noloc.xlsx", BASE_MAPPING)
    body = resp.json()
    assert body["summary"]["customers_created"] == 0


# -- Import: create/update/rollback/partial ----------------------------------------

async def test_commit_creates_customer_territory_invoice(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("CREATE")
    territory_name = _unique("CreateTerritory")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["territory_names"].append(territory_name)
    import_cleanup["filenames"].append("create.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "Create Test Outlet", "1 Main St", "+919800000010", "12.5", "77.5", "", territory_name, "INV-CR1", "2026-06-01", "12345", "Usha", "", "", "", ""]])
    v = await _validate(client, admin_headers, data, "create.xlsx", BASE_MAPPING)
    batch_id = v.json()["id"]
    c = await client.post(f"/api/v1/imports/{batch_id}/commit", headers=admin_headers)
    assert c.status_code == 200, c.text
    assert c.json()["status"] == "COMMITTED"

    from tests.integration.conftest import db_cursor as _dbc
    with _dbc() as cur:
        cur.execute("SELECT name, territory_id FROM customers WHERE outlet_code = %s", (code,))
        row = cur.fetchone()
        assert row is not None
        assert row["name"] == "Create Test Outlet"
        cur.execute("SELECT amount, source FROM invoices WHERE invoice_number = 'INV-CR1'")
        inv = cur.fetchone()
        assert inv is not None
        assert str(inv["amount"]) == "12345.00"
        assert inv["source"] == "EXCEL_IMPORT"


async def test_reimport_same_file_updates_not_duplicates(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("REIMPORT")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("reimport.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "Reimport Outlet", "1 Main St", "+919800000011", "12.5", "77.5", "", "", "INV-RE1", "2026-06-01", "1000", "Usha", "", "", "", ""]])

    v1 = await _validate(client, admin_headers, data, "reimport.xlsx", BASE_MAPPING)
    b1 = v1.json()["id"]
    r1 = await client.post(f"/api/v1/imports/{b1}/commit", headers=admin_headers)
    assert r1.json()["summary"]["customers_created"] == 1

    v2 = await _validate(client, admin_headers, data, "reimport.xlsx", BASE_MAPPING)
    b2 = v2.json()["id"]
    body2 = v2.json()
    assert body2["summary"]["customers_created"] == 0
    assert body2["summary"]["customers_updated"] == 1
    assert body2["summary"]["invoices_created"] == 0
    assert body2["summary"]["invoices_updated"] == 1
    r2 = await client.post(f"/api/v1/imports/{b2}/commit", headers=admin_headers)
    assert r2.status_code == 200

    with db_cursor() as cur:
        cur.execute("SELECT COUNT(*) as n FROM customers WHERE outlet_code = %s", (code,))
        assert cur.fetchone()["n"] == 1
        cur.execute("SELECT COUNT(*) as n FROM invoices WHERE invoice_number = 'INV-RE1'")
        assert cur.fetchone()["n"] == 1


async def test_partial_invalid_rows_do_not_block_valid_rows(client: AsyncClient, admin_headers, import_cleanup):
    good_code = _unique("GOOD")
    bad_code = _unique("BAD")
    import_cleanup["outlet_codes"] += [good_code, bad_code]
    import_cleanup["filenames"].append("partial.xlsx")
    rows = [
        [good_code, "Good Outlet", "Addr", "+919800000012", "12.0", "77.0", "", "", "INV-GOOD1", "2026-06-01", "1000", "Usha", "", "", "", ""],
        # bad_code's outlet itself is unresolvable (no coordinates for a
        # brand-new outlet) - unlike an invoice-only problem, this must
        # block the customer too, since there's nothing safe to create.
        [bad_code, "Bad Outlet", "Addr", "+919800000013", "", "", "", "", "INV-BAD1", "2026-06-01", "1000", "Usha", "", "", "", ""],
    ]
    data = _xlsx(FULL_HEADERS, rows)
    v = await _validate(client, admin_headers, data, "partial.xlsx", BASE_MAPPING)
    body = v.json()
    assert body["summary"]["customers_created"] == 1
    batch_id = body["id"]
    c = await client.post(f"/api/v1/imports/{batch_id}/commit", headers=admin_headers)
    assert c.status_code == 200

    with db_cursor() as cur:
        cur.execute("SELECT COUNT(*) as n FROM customers WHERE outlet_code = %s", (good_code,))
        assert cur.fetchone()["n"] == 1
        cur.execute("SELECT COUNT(*) as n FROM customers WHERE outlet_code = %s", (bad_code,))
        assert cur.fetchone()["n"] == 0


async def test_invoice_only_error_does_not_block_the_outlets_own_creation(
    client: AsyncClient, admin_headers, import_cleanup
):
    """A row's outlet data can be perfectly valid even if that row's invoice date is garbage - the outlet must still be created."""
    code = _unique("INVOK")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("invok.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "Invoice Only Bad Outlet", "Addr", "+919800000099", "12.0", "77.0", "", "", "INV-BAD9", "not-a-date", "1000", "Usha", "", "", "", ""]])
    v = await _validate(client, admin_headers, data, "invok.xlsx", BASE_MAPPING)
    body = v.json()
    assert body["rows_error"] == 1
    assert body["summary"]["customers_created"] == 1
    batch_id = body["id"]
    c = await client.post(f"/api/v1/imports/{batch_id}/commit", headers=admin_headers)
    assert c.status_code == 200

    with db_cursor() as cur:
        cur.execute("SELECT COUNT(*) as n FROM customers WHERE outlet_code = %s", (code,))
        assert cur.fetchone()["n"] == 1
        cur.execute("SELECT COUNT(*) as n FROM invoices WHERE invoice_number = 'INV-BAD9'")
        assert cur.fetchone()["n"] == 0


async def test_cannot_commit_a_batch_twice(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("TWICE")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("twice.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "Twice Outlet", "Addr", "+919800000014", "12.0", "77.0", "", "", "INV-TW1", "2026-06-01", "1000", "Usha", "", "", "", ""]])
    v = await _validate(client, admin_headers, data, "twice.xlsx", BASE_MAPPING)
    batch_id = v.json()["id"]
    r1 = await client.post(f"/api/v1/imports/{batch_id}/commit", headers=admin_headers)
    assert r1.status_code == 200
    r2 = await client.post(f"/api/v1/imports/{batch_id}/commit", headers=admin_headers)
    assert r2.status_code == 409


# -- History / error report -------------------------------------------------------

async def test_import_history_lists_batches(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("HIST")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("history_test.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "History Outlet", "Addr", "+919800000015", "12.0", "77.0", "", "", "INV-H1", "2026-06-01", "1000", "Usha", "", "", "", ""]])
    v = await _validate(client, admin_headers, data, "history_test.xlsx", BASE_MAPPING)
    batch_id = v.json()["id"]

    resp = await client.get("/api/v1/imports", headers=admin_headers)
    assert resp.status_code == 200
    assert any(b["id"] == batch_id for b in resp.json())

    detail = await client.get(f"/api/v1/imports/{batch_id}", headers=admin_headers)
    assert detail.status_code == 200
    assert detail.json()["filename"] == "history_test.xlsx"


async def test_error_report_csv_download(client: AsyncClient, admin_headers, import_cleanup):
    code = _unique("CSVERR")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("csverr.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "CSV Error Outlet", "Addr", "+919800000016", "12.0", "77.0", "", "", "INV-CE1", "bad-date", "1000", "Usha", "", "", "", ""]])
    v = await _validate(client, admin_headers, data, "csverr.xlsx", BASE_MAPPING)
    batch_id = v.json()["id"]
    resp = await client.get(f"/api/v1/imports/{batch_id}/errors.csv", headers=admin_headers)
    assert resp.status_code == 200
    assert "Row" in resp.text
    assert "invoice_date" in resp.text


# -- Authorization ------------------------------------------------------------------

async def test_employee_cannot_preview(client: AsyncClient, employee_headers):
    data = _xlsx(FULL_HEADERS, [["A"] * len(FULL_HEADERS)])
    resp = await client.post(
        "/api/v1/imports/preview",
        files={"file": ("f.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=employee_headers,
    )
    assert resp.status_code == 403


async def test_employee_cannot_validate(client: AsyncClient, employee_headers):
    data = _xlsx(FULL_HEADERS, [["A"] * len(FULL_HEADERS)])
    resp = await _validate(client, employee_headers, data, "f.xlsx", BASE_MAPPING)
    assert resp.status_code == 403


async def test_employee_cannot_commit(client: AsyncClient, admin_headers, employee_headers, import_cleanup):
    code = _unique("AUTHZ")
    import_cleanup["outlet_codes"].append(code)
    import_cleanup["filenames"].append("authz.xlsx")
    data = _xlsx(FULL_HEADERS, [[code, "Authz Outlet", "Addr", "+919800000017", "12.0", "77.0", "", "", "INV-AZ1", "2026-06-01", "1000", "Usha", "", "", "", ""]])
    v = await _validate(client, admin_headers, data, "authz.xlsx", BASE_MAPPING)
    batch_id = v.json()["id"]
    resp = await client.post(f"/api/v1/imports/{batch_id}/commit", headers=employee_headers)
    assert resp.status_code == 403


async def test_employee_cannot_see_import_history(client: AsyncClient, employee_headers):
    resp = await client.get("/api/v1/imports", headers=employee_headers)
    assert resp.status_code == 403


async def test_unauthenticated_cannot_preview(client: AsyncClient):
    data = _xlsx(FULL_HEADERS, [["A"] * len(FULL_HEADERS)])
    resp = await client.post(
        "/api/v1/imports/preview",
        files={"file": ("f.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 401
